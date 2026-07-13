import os
import io
import json
import time
import shutil
import socket
import secrets
import logging
import calendar
import threading
import sqlite3
import qrcode
import urllib.request
import urllib.error
from urllib.parse import urlencode
from datetime import datetime, date, timedelta, time as dt_time
from flask import (Flask, render_template, request, redirect, url_for,
                   flash, send_file, jsonify, abort, session)
from database import get_db, init_db
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.text import RichText
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.chart.layout import Layout, ManualLayout
from werkzeug.utils import secure_filename, safe_join
from werkzeug.security import generate_password_hash, check_password_hash
from cryptography.fernet import Fernet, InvalidToken

app = Flask(__name__)

BASE_DIR    = os.path.dirname(__file__)
LOG_PATH    = os.path.join(BASE_DIR, 'server.log')

# ─── 로깅 (보안 이벤트 등 서버 로그를 파일로 보관) ──────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_PATH, encoding='utf-8'),
        logging.StreamHandler(),
    ]
)

# ─── 세션 서명용 비밀키 (최초 실행 시 랜덤 생성 후 파일로 보관) ──
SECRET_KEY_PATH = os.path.join(BASE_DIR, 'secret_key.txt')
def _load_or_create_secret_key():
    if os.path.exists(SECRET_KEY_PATH):
        with open(SECRET_KEY_PATH, encoding='utf-8') as f:
            key = f.read().strip()
            if key:
                return key
    key = secrets.token_hex(32)
    with open(SECRET_KEY_PATH, 'w', encoding='utf-8') as f:
        f.write(key)
    return key

app.secret_key = _load_or_create_secret_key()
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 요청당 최대 50MB (사진 최대 10장 대비)
app.permanent_session_lifetime = timedelta(minutes=60)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# ─── CSRF 보호 (세션 기반 토큰) ──────────────────────────
@app.context_processor
def _inject_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(16)
    return dict(csrf_token=lambda: session['csrf_token'])

@app.before_request
def _csrf_protect():
    if request.method == 'POST':
        token = session.get('csrf_token')
        submitted = request.form.get('csrf_token') or request.headers.get('X-CSRFToken')
        if not token or submitted != token:
            flash('보안 토큰이 만료되었습니다. 다시 시도해주세요.', 'danger')
            return redirect(request.referrer or url_for('index'))

@app.context_processor
def _inject_actor_workers():
    db = get_db()
    names = get_workers(db)
    db.close()
    return dict(actor_workers=names)

# ─── 보안 응답 헤더 ──────────────────────────────────────
@app.after_request
def _set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "img-src 'self' data:; "
        "connect-src 'self'"
    )
    response.headers.pop('Server', None)
    return response

# ─── 로그인 벽 (QR 고장등록 경로는 예외) ─────────────────
LOGIN_EXEMPT_ENDPOINTS = {'login', 'static', 'fault_register', 'fault_select', 'kakao_callback', 'worker_delete', 'service_worker'}

@app.before_request
def _require_login():
    if request.endpoint is None or request.endpoint in LOGIN_EXEMPT_ENDPOINTS:
        return
    if not session.get('edit_authorized'):
        return redirect(url_for('login', next=request.full_path.rstrip('?')))

QR_DIR      = os.path.join(BASE_DIR, 'qr_codes')
UPLOAD_DIR  = os.path.join(BASE_DIR, 'uploads')
FAULT_DIR   = os.path.join(UPLOAD_DIR, 'fault')
ACTION_DIR  = os.path.join(UPLOAD_DIR, 'action')
PHOTOS_DIR  = os.path.join(UPLOAD_DIR, 'photos')

os.makedirs(PHOTOS_DIR, exist_ok=True)

ALLOWED_EXT = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

SECTIONS = ['Section 1', 'Section 2', 'Section 3',
            'Section 4.1', 'Section 4.2', 'Section 4.3', 'Section 4.4']

# 가동정지 사유 목록. '퇴근'만 업무시간 외로 보고 가동대상시간 자체에서 제외한다
# (점심시간 등 휴식시간과 동일하게 취급). 그 외 사유(설비점검/자재대기/교체/기타)는
# 업무시간 중 실제로 설비가 멈춘 것이므로 고장과 동일하게 설비가동률에 반영(차감)된다.
PAUSE_REASONS = ['퇴근', '설비점검', '자재대기', '교체', '기타']
PAUSE_NONWORKING_REASONS = {'퇴근'}

SYMPTOMS = [
    '센서 감지불량', '실린더 동작불량', '에어누설 (Air Leak)',
    '배큠흡착불량', '히터 불량', '부품 파손', '체인 불량',
    '베어링 소음', '벨트 마모', '체결부 풀림',
    '제품 걸림 (JAM)', 'HMI 알람', 'PLC 통신이상', '누액 (Leak)',
    '밸브 동작불량', '공정 타임아웃 (진행 안됨)', '카메라 인식오류', '도어 개폐 불량',
    '기타'
]

PARTS = ['센서', '실린더', '히터', '베어링', '벨트', '멤브레인 시트', '포팅스타']

GRADES = [('A', 'A : 생산정지'), ('B', 'B : 품질영향'), ('C', 'C : 경미고장'), ('D', 'D : 기타')]
STATUSES = ['미조치', '진행중', '대기', '완료']

# ─── 관리자 비밀번호 (해시로 저장, 원래 값은 synopex) ──────────
ADMIN_CONFIG_PATH = os.path.join(BASE_DIR, 'admin_config.json')
_DEFAULT_ADMIN_PASSWORD = 'synopex'

def _load_admin_password_hash():
    try:
        with open(ADMIN_CONFIG_PATH, encoding='utf-8') as f:
            return json.load(f)['password_hash']
    except Exception:
        h = generate_password_hash(_DEFAULT_ADMIN_PASSWORD)
        with open(ADMIN_CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump({'password_hash': h}, f)
        logging.warning(
            'admin_config.json이 없어 기본 비밀번호(%s)로 새로 생성했습니다. 로그인 후 반드시 비밀번호를 변경하세요.',
            _DEFAULT_ADMIN_PASSWORD)
        return h

def check_admin_password(pw):
    return check_password_hash(_load_admin_password_hash(), pw or '')

def set_admin_password(new_password):
    h = generate_password_hash(new_password)
    with open(ADMIN_CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump({'password_hash': h}, f)


# ─── 로그인 실패 잠금 (brute-force 방지, 서비스 재시작에도 유지) ──
_LOGIN_MAX_ATTEMPTS    = 5
_LOGIN_LOCKOUT_SECONDS = 300
LOGIN_ATTEMPTS_PATH = os.path.join(BASE_DIR, 'login_attempts.json')

def _load_login_attempts():
    try:
        with open(LOGIN_ATTEMPTS_PATH, encoding='utf-8') as f:
            data = json.load(f)
        now = time.time()
        return {ip: info for ip, info in data.items() if info.get('locked_until', 0) > now}
    except Exception:
        return {}

def _save_login_attempts():
    try:
        with open(LOGIN_ATTEMPTS_PATH, 'w', encoding='utf-8') as f:
            json.dump(_login_attempts, f)
    except Exception:
        logging.exception('login_attempts.json 저장 실패')

_login_attempts = _load_login_attempts()  # ip -> {'count': int, 'locked_until': float}

def _login_lock_remaining(ip):
    info = _login_attempts.get(ip)
    if not info:
        return 0
    remaining = info['locked_until'] - time.time()
    if remaining > 0:
        logging.warning('잠긴 IP의 접근 시도 차단: %s (남은 잠금시간 %d초)', ip, int(remaining))
    return max(0, remaining)

def _register_login_failure(ip):
    info = _login_attempts.setdefault(ip, {'count': 0, 'locked_until': 0})
    info['count'] += 1
    logging.warning('로그인 비밀번호 실패: %s (%d/%d회)', ip, info['count'], _LOGIN_MAX_ATTEMPTS)
    if info['count'] >= _LOGIN_MAX_ATTEMPTS:
        info['locked_until'] = time.time() + _LOGIN_LOCKOUT_SECONDS
        info['count'] = 0
        logging.warning('brute-force 의심으로 IP 잠금: %s (%d초간)', ip, _LOGIN_LOCKOUT_SECONDS)
        _notify_login_lockout(ip)
    _save_login_attempts()

def _notify_login_lockout(ip):
    if not _kakao_connected():
        return
    text = (f'🔒 보안 경고 - 로그인 잠금\n'
            f'IP {ip}에서 비밀번호를 {_LOGIN_MAX_ATTEMPTS}회 연속 틀려 {_LOGIN_LOCKOUT_SECONDS // 60}분간 접속이 차단되었습니다.\n'
            f'본인이 아니라면 무단 접근 시도일 수 있습니다.')
    try:
        _kakao_send_to_all(text, f'http://{get_local_ip()}:5001/login')
    except Exception:
        logging.exception('_notify_login_lockout failed')

def _register_login_success(ip):
    if _login_attempts.pop(ip, None) is not None:
        _save_login_attempts()


# ─── 카카오톡 알림 연동 (A등급 고장 발생 시) ──────────────────
KAKAO_CONFIG_PATH   = os.path.join(BASE_DIR, 'kakao_config.json')
KAKAO_ENCRYPT_KEY_PATH = os.path.join(BASE_DIR, 'kakao_encrypt.key')
KAKAO_AUTHORIZE_URL = 'https://kauth.kakao.com/oauth/authorize'
KAKAO_TOKEN_URL     = 'https://kauth.kakao.com/oauth/token'
KAKAO_SEND_URL      = 'https://kapi.kakao.com/v2/api/talk/memo/default/send'
KAKAO_USER_ME_URL   = 'https://kapi.kakao.com/v2/user/me'

# ─── 알림 유형별 on/off 설정 (사람마다 원하는 알림만 받고 싶을 수 있어서 분리) ──
NOTIFY_SETTINGS_PATH = os.path.join(BASE_DIR, 'notify_settings.json')
NOTIFY_TYPES = {'fault': '고장 발생 알림', 'section_end': '섹션 가동종료 알림'}

def _load_notify_settings():
    try:
        with open(NOTIFY_SETTINGS_PATH, encoding='utf-8') as f:
            cfg = json.load(f)
    except Exception:
        cfg = {}
    return {k: cfg.get(k, True) for k in NOTIFY_TYPES}  # 기본값: 전부 켜짐(기존 동작 유지)

def _save_notify_settings(cfg):
    with open(NOTIFY_SETTINGS_PATH, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False)

def _notify_type_enabled(key):
    return _load_notify_settings().get(key, True)

try:
    import win32crypt
    _DPAPI_LOCAL_MACHINE = 4  # CRYPTPROTECT_LOCAL_MACHINE - 이 기기 밖으로 파일이 유출돼도 복호화 불가
    _DPAPI_AVAILABLE = True
except ImportError:
    _DPAPI_AVAILABLE = False

def _dpapi_protect(data: bytes) -> bytes:
    return win32crypt.CryptProtectData(data, 'kakao_encrypt_key', None, None, None, _DPAPI_LOCAL_MACHINE)

def _dpapi_unprotect(data: bytes) -> bytes:
    return win32crypt.CryptUnprotectData(data, None, None, None, _DPAPI_LOCAL_MACHINE)[1]

def _save_kakao_key(key: bytes):
    with open(KAKAO_ENCRYPT_KEY_PATH, 'wb') as f:
        f.write(_dpapi_protect(key) if _DPAPI_AVAILABLE else key)

def _load_or_create_kakao_key():
    if os.path.exists(KAKAO_ENCRYPT_KEY_PATH):
        with open(KAKAO_ENCRYPT_KEY_PATH, 'rb') as f:
            raw = f.read().strip()
        if raw:
            if _DPAPI_AVAILABLE:
                try:
                    return _dpapi_unprotect(raw)
                except Exception:
                    pass
            # DPAPI 보호 전(레거시) 평문 키였을 수 있음 - 유효하면 그대로 쓰고 다음 저장 시 DPAPI로 이전
            try:
                Fernet(raw)
                _save_kakao_key(raw)
                return raw
            except Exception:
                logging.warning('kakao_encrypt.key를 읽지 못했습니다 - 새 키를 생성합니다.')
    key = Fernet.generate_key()
    _save_kakao_key(key)
    return key

_kakao_fernet = Fernet(_load_or_create_kakao_key())
_KAKAO_ENC_PREFIX = 'enc:'
_KAKAO_SECRET_FIELDS = ('client_secret',)
_KAKAO_ACCOUNT_SECRET_FIELDS = ('access_token', 'refresh_token')

def _kakao_encrypt_value(v):
    if not v or not isinstance(v, str) or v.startswith(_KAKAO_ENC_PREFIX):
        return v
    return _KAKAO_ENC_PREFIX + _kakao_fernet.encrypt(v.encode('utf-8')).decode('ascii')

def _kakao_decrypt_value(v):
    if not v or not isinstance(v, str) or not v.startswith(_KAKAO_ENC_PREFIX):
        return v  # 평문(암호화 적용 이전 값) 그대로 반환 - 다음 저장 시 자동 암호화됨
    try:
        return _kakao_fernet.decrypt(v[len(_KAKAO_ENC_PREFIX):].encode('ascii')).decode('utf-8')
    except InvalidToken:
        logging.warning('kakao_config.json 복호화 실패 - 키 파일이 바뀌었을 수 있습니다.')
        return v

def _kakao_load():
    try:
        with open(KAKAO_CONFIG_PATH, encoding='utf-8') as f:
            cfg = json.load(f)
    except Exception:
        cfg = {}
    cfg.setdefault('accounts', [])
    for field in _KAKAO_SECRET_FIELDS:
        if field in cfg:
            cfg[field] = _kakao_decrypt_value(cfg[field])
    for account in cfg['accounts']:
        for field in _KAKAO_ACCOUNT_SECRET_FIELDS:
            if field in account:
                account[field] = _kakao_decrypt_value(account[field])
    return cfg

def _kakao_save(cfg):
    out = dict(cfg)
    for field in _KAKAO_SECRET_FIELDS:
        if field in out:
            out[field] = _kakao_encrypt_value(out[field])
    out['accounts'] = []
    for account in cfg.get('accounts', []):
        acc = dict(account)
        for field in _KAKAO_ACCOUNT_SECRET_FIELDS:
            if field in acc:
                acc[field] = _kakao_encrypt_value(acc[field])
        out['accounts'].append(acc)
    with open(KAKAO_CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

def _kakao_connected():
    return bool(_kakao_load().get('accounts'))

def _kakao_authorize_url():
    cfg = _kakao_load()
    params = {
        'client_id': cfg.get('rest_api_key', ''),
        'redirect_uri': cfg.get('redirect_uri', ''),
        'response_type': 'code',
        'scope': 'talk_message',
    }
    return KAKAO_AUTHORIZE_URL + '?' + urlencode(params)

def _kakao_post_form(url, data, headers=None):
    body = urlencode(data).encode('utf-8')
    req = urllib.request.Request(url, data=body, method='POST')
    req.add_header('Content-Type', 'application/x-www-form-urlencoded;charset=utf-8')
    for k, v in (headers or {}).items():
        req.add_header(k, v)
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        try:
            return json.loads(e.read().decode('utf-8'))
        except Exception:
            return {'error': f'http_{e.code}'}

def _kakao_get_profile(access_token):
    req = urllib.request.Request(KAKAO_USER_ME_URL, method='GET')
    req.add_header('Authorization', f'Bearer {access_token}')
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
        kakao_id = str(data.get('id', ''))
        nickname = (data.get('properties') or {}).get('nickname', '') or f'사용자 {kakao_id[-4:]}'
        return kakao_id, nickname
    except Exception:
        logging.exception('kakao profile fetch failed')
        return None, None

def _kakao_exchange_code(code):
    cfg = _kakao_load()
    data = {
        'grant_type': 'authorization_code',
        'client_id': cfg.get('rest_api_key', ''),
        'redirect_uri': cfg.get('redirect_uri', ''),
        'code': code,
    }
    if cfg.get('client_secret'):
        data['client_secret'] = cfg['client_secret']
    result = _kakao_post_form(KAKAO_TOKEN_URL, data)
    if 'access_token' not in result:
        return result, None

    kakao_id, nickname = _kakao_get_profile(result['access_token'])
    account = {
        'kakao_id': kakao_id or f'user_{int(time.time())}',
        'nickname': nickname or '이름 없음',
        'access_token': result['access_token'],
        'refresh_token': result.get('refresh_token'),
        'token_expires_at': time.time() + result.get('expires_in', 21599) - 60,
    }
    accounts = [a for a in cfg['accounts'] if a.get('kakao_id') != account['kakao_id']]
    accounts.append(account)
    cfg['accounts'] = accounts
    _kakao_save(cfg)
    return result, account

def _kakao_refresh_account(account):
    cfg = _kakao_load()
    if not account.get('refresh_token'):
        return None
    data = {
        'grant_type': 'refresh_token',
        'client_id': cfg.get('rest_api_key', ''),
        'refresh_token': account['refresh_token'],
    }
    if cfg.get('client_secret'):
        data['client_secret'] = cfg['client_secret']
    result = _kakao_post_form(KAKAO_TOKEN_URL, data)
    if 'access_token' not in result:
        logging.warning('kakao refresh failed for %s: %s', account.get('nickname'), result)
        return None
    account['access_token'] = result['access_token']
    if 'refresh_token' in result:
        account['refresh_token'] = result['refresh_token']
    account['token_expires_at'] = time.time() + result.get('expires_in', 21599) - 60
    for i, a in enumerate(cfg['accounts']):
        if a.get('kakao_id') == account.get('kakao_id'):
            cfg['accounts'][i] = account
    _kakao_save(cfg)
    return account['access_token']

def _kakao_valid_token_for(account):
    if time.time() >= (account.get('token_expires_at') or 0):
        return _kakao_refresh_account(account)
    return account.get('access_token')

def _kakao_remove_account(kakao_id):
    cfg = _kakao_load()
    cfg['accounts'] = [a for a in cfg['accounts'] if a.get('kakao_id') != kakao_id]
    _kakao_save(cfg)

def _kakao_send_to_account(account, text, link_url):
    token = _kakao_valid_token_for(account)
    if not token:
        return False, '토큰 갱신 실패'
    template = {
        'object_type': 'text',
        'text': text,
        'link': {'web_url': link_url, 'mobile_web_url': link_url},
    }
    result = _kakao_post_form(
        KAKAO_SEND_URL,
        {'template_object': json.dumps(template, ensure_ascii=False)},
        headers={'Authorization': f'Bearer {token}'}
    )
    if result.get('result_code') == 0:
        return True, result
    logging.warning('kakao send failed for %s: %s', account.get('nickname'), result)
    return False, result

def _kakao_send_to_all(text, link_url):
    cfg = _kakao_load()
    results = []
    for account in cfg.get('accounts', []):
        ok, result = _kakao_send_to_account(account, text, link_url)
        results.append((account.get('nickname'), ok, result))
    return results

GRADE_NOTIFY_LABEL = {
    'A': ('🚨', 'A등급 (생산정지)'),
    'B': ('⚠️', 'B등급 (품질영향)'),
    'C': ('🔧', 'C등급 (경미고장)'),
}

def notify_fault_by_grade(grade, fault_id, eq_number, eq_name, symptom, status=None, worker=None, action_detail=None):
    if grade not in GRADE_NOTIFY_LABEL or not _kakao_connected() or not _notify_type_enabled('fault'):
        return
    emoji, label = GRADE_NOTIFY_LABEL[grade]
    ip = get_local_ip()
    link_url = f'http://{ip}:5001/fault/{fault_id}'
    text = (f'{emoji} {label} 고장 발생\n'
            f'설비: {eq_number} {eq_name}\n'
            f'증상: {symptom or "-"}\n'
            f'처리상태: {status or "미조치"}\n'
            f'작업자: {worker or "-"}\n'
            f'조치내용: {action_detail or "-"}\n'
            f'상세보기에서 확인해주세요.')
    try:
        _kakao_send_to_all(text, link_url)
    except Exception:
        logging.exception('notify_fault_by_grade failed')


def notify_section_end(lot, section, stats):
    if not _kakao_connected() or not _notify_type_enabled('section_end'):
        return
    rate = stats['uptime_rate']
    if rate >= 95:
        emoji, label = '✅', '정상'
    elif rate >= 85:
        emoji, label = '⚠️', '주의'
    else:
        emoji, label = '🔴', '저조'
    avail_m = stats['available_seconds'] // 60
    fault_m = stats['fault_seconds'] // 60
    pause_m = stats['pause_downtime_seconds'] // 60
    down_m = fault_m + pause_m
    ip = get_local_ip()
    link_url = f'http://{ip}:5001/production/{lot["id"]}'
    text = (f'{emoji} 설비가동 섹션 종료 ({label})\n'
            f'LOT: {lot["model_name"]} / {lot["lot_number"]}\n'
            f'섹션: {section}\n'
            f'가동률: {rate}%\n'
            f'가동대상시간: {avail_m}분\n'
            f'총 설비정지: {down_m}분 (고장 {fault_m}분 {stats["fault_count"]}건 · 정지 {pause_m}분 {stats["pause_count"]}건)\n'
            f'상세보기에서 확인해주세요.')
    try:
        _kakao_send_to_all(text, link_url)
    except Exception:
        logging.exception('notify_section_end failed')


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


PHOTO_MAX_DIM = 1600
PHOTO_JPEG_QUALITY = 80

def save_and_compress_image(file_storage, dest_path):
    img = Image.open(file_storage.stream)
    img = img.convert('RGB') if img.mode != 'RGB' else img
    w, h = img.size
    if max(w, h) > PHOTO_MAX_DIM:
        scale = PHOTO_MAX_DIM / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    img.save(dest_path, 'JPEG', quality=PHOTO_JPEG_QUALITY, optimize=True)


def last_n_months(n=12):
    """오늘 기준 최근 n개월을 'YYYY-MM' 형식으로, 최신순으로 반환 (데이터 유무와 무관)."""
    months = []
    today = date.today()
    y, m = today.year, today.month
    for i in range(n):
        mm = m - i
        yy = y
        while mm <= 0:
            mm += 12
            yy -= 1
        months.append(f'{yy:04d}-{mm:02d}')
    return months


def get_workers(db):
    rows = db.execute("SELECT name FROM workers ORDER BY name").fetchall()
    return [r['name'] for r in rows]


def register_worker(db, name):
    if name:
        db.execute("INSERT OR IGNORE INTO workers (name) VALUES (?)", (name,))


# ─── 생산 LOT / 설비가동률 ────────────────────────────────
def get_active_lot_section(db, section):
    """해당 섹션에서 현재 진행중(가동 시작을 눌렀지만 종료 안 한) 구간을 반환."""
    return db.execute(
        "SELECT * FROM production_lot_sections WHERE section=? AND status='진행중' ORDER BY started_at DESC LIMIT 1",
        (section,)
    ).fetchone()


def _parse_dt(s):
    """'YYYY-MM-DD HH:MM:SS' 형식이 기본이지만, 생산LOT 기능 도입 이전에 날짜만(시간 없이)
    저장된 예전 고장이력(completed_at)도 있으므로 'YYYY-MM-DD'는 자정으로 보고 폴백 파싱한다."""
    try:
        return datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        return datetime.strptime(s, '%Y-%m-%d')


def _overlap_seconds(a_start, a_end, b_start, b_end):
    """[a_start,a_end) 구간과 [b_start,b_end) 구간이 겹치는 초 (안 겹치면 0)."""
    latest_start = max(a_start, b_start)
    earliest_end = min(a_end, b_end)
    return max(0, (earliest_end - latest_start).total_seconds())


def _get_section_pauses(db, lot_section_id, win_start, win_end):
    """섹션의 가동정지 기록을 win_start~win_end로 클리핑해서 반환.
    'nonworking'=True(퇴근)는 업무시간 외로 보고 가동대상시간에서 제외되고,
    False(설비점검/자재대기/교체/기타)는 업무시간 중 실제 다운타임으로 설비가동률에 반영된다."""
    pauses = db.execute(
        "SELECT * FROM production_lot_section_pauses WHERE lot_section_id=? ORDER BY paused_at",
        (lot_section_id,)
    ).fetchall()
    result = []
    for p in pauses:
        s = max(win_start, _parse_dt(p['paused_at']))
        e = min(win_end, _parse_dt(p['resumed_at']) if p['resumed_at'] else win_end)
        if e > s:
            reason = p['reason'] or '퇴근'
            result.append({
                'start': s, 'end': e, 'reason': reason,
                'label': f'가동정지({reason})',
                'nonworking': reason in PAUSE_NONWORKING_REASONS,
            })
    return result


def _excluded_intervals(db, lot_section_id, win_start, win_end):
    """가동대상시간(분모)에서 빠지는 구간(반복되는 휴식시간 + '퇴근' 사유의 가동정지)을
    시간순으로 병합해서 반환. 겹치는 구간은 하나로 합쳐서 이중으로 차감되지 않도록 한다.
    '잔업'은 휴식시간이 아니라 근무시간이므로 여기서 제외 대상에서 뺀다 — 잔업 중 실제로
    가동한 시간은 그대로 가동대상시간에 포함되어 가동률 계산에 반영된다."""
    schedules = db.execute(
        "SELECT name, start_time, end_time FROM break_schedules WHERE name != '잔업'"
    ).fetchall()
    raw = []
    day = win_start.date()
    end_day = win_end.date()
    while day <= end_day:
        for sch in schedules:
            try:
                sh, sm = (int(x) for x in sch['start_time'].split(':'))
                eh, em = (int(x) for x in sch['end_time'].split(':'))
                b_start = datetime.combine(day, dt_time(sh, sm))
                b_end = datetime.combine(day, dt_time(eh, em))
            except (ValueError, TypeError):
                continue
            if b_end <= b_start:
                continue
            s, e = max(win_start, b_start), min(win_end, b_end)
            if e > s:
                raw.append({'start': s, 'end': e, 'label': sch['name'], 'kind': 'break'})
        day += timedelta(days=1)

    for p in _get_section_pauses(db, lot_section_id, win_start, win_end):
        if p['nonworking']:
            raw.append({'start': p['start'], 'end': p['end'], 'label': p['label'], 'kind': 'pause'})

    # 수동 가동정지가 우선순위 높음 (겹치면 정지 사유 라벨을 그대로 유지)
    raw.sort(key=lambda x: (x['start'], x['kind'] != 'pause'))
    merged = []
    for iv in raw:
        if merged and iv['start'] <= merged[-1]['end']:
            if iv['end'] > merged[-1]['end']:
                merged[-1]['end'] = iv['end']
            if merged[-1]['kind'] != 'pause' and iv['kind'] == 'pause':
                merged[-1]['label'] = iv['label']
                merged[-1]['kind'] = 'pause'
        else:
            merged.append(dict(iv))
    return merged


def _downtime_pause_intervals(db, lot_section_id, win_start, win_end):
    """가동대상시간(분모) 안에서 실제 다운타임으로 카운트되는 가동정지 구간
    (퇴근 외 사유). 고장과 마찬가지로 설비가동률에 반영(차감)된다."""
    return [p for p in _get_section_pauses(db, lot_section_id, win_start, win_end) if not p['nonworking']]


def compute_section_stats(db, lot_section_id, lot_section_row=None):
    """섹션 가동구간 하나(시작~종료)의 가동대상시간/고장시간/가동률 계산.
    아직 진행 중(종료 안 함)이면 현재 시각까지로 계산한 미리보기 값을 반환한다."""
    ls = lot_section_row or db.execute(
        "SELECT * FROM production_lot_sections WHERE id=?", (lot_section_id,)
    ).fetchone()
    if not ls:
        return None

    win_start = _parse_dt(ls['started_at'])
    win_end = _parse_dt(ls['ended_at']) if ls['ended_at'] else datetime.now()

    total_seconds = max(0.0, (win_end - win_start).total_seconds())
    excluded = _excluded_intervals(db, lot_section_id, win_start, win_end)
    excluded_seconds = sum((iv['end'] - iv['start']).total_seconds() for iv in excluded)
    available_seconds = max(0.0, total_seconds - excluded_seconds)

    faults = db.execute("""
        SELECT occurred_at, completed_at, symptom, cause, action_detail, worker
        FROM fault_history WHERE lot_section_id = ?
        ORDER BY occurred_at
    """, (lot_section_id,)).fetchall()

    fault_seconds = 0.0
    fault_rows = []
    for f in faults:
        occ = _parse_dt(f['occurred_at'])
        comp = _parse_dt(f['completed_at']) if f['completed_at'] else win_end
        comp = min(comp, win_end)
        duration = max(0.0, (comp - occ).total_seconds())
        fault_seconds += duration
        fault_rows.append({
            'occurred_at': f['occurred_at'],
            'completed_at': f['completed_at'],
            'duration_minutes': int(duration // 60),
            'symptom': f['symptom'],
            'cause': f['cause'],
            'action_detail': f['action_detail'],
            'worker': f['worker'],
            'ongoing': f['completed_at'] is None,
        })

    downtime_pauses = _downtime_pause_intervals(db, lot_section_id, win_start, win_end)
    pause_downtime_seconds = 0.0
    pause_rows = []
    for p in downtime_pauses:
        duration = (p['end'] - p['start']).total_seconds()
        pause_downtime_seconds += duration
        pause_rows.append({
            'paused_at': p['start'].strftime('%Y-%m-%d %H:%M:%S'),
            'resumed_at': p['end'].strftime('%Y-%m-%d %H:%M:%S') if p['end'] < win_end else None,
            'duration_minutes': int(duration // 60),
            'reason': p['reason'],
            'ongoing': p['end'] >= win_end,
        })

    downtime_total = fault_seconds + pause_downtime_seconds
    if available_seconds > 0:
        uptime_rate = max(0.0, min(100.0, (available_seconds - downtime_total) / available_seconds * 100))
    else:
        uptime_rate = 100.0

    return {
        'section': ls['section'],
        'available_seconds': int(available_seconds),
        'fault_seconds': int(fault_seconds),
        'fault_count': len(faults),
        'pause_downtime_seconds': int(pause_downtime_seconds),
        'pause_count': len(pause_rows),
        'uptime_rate': round(uptime_rate, 1),
        'faults': fault_rows,
        'pauses': pause_rows,
    }


def get_section_dashboard_status(db):
    """대시보드용: 섹션(SECTIONS) 각각의 현재 가동 상태와 실시간 가동률.
    대기 중인 섹션은, 진행중인 LOT이 정확히 하나뿐이고 그 LOT에서 아직 안 쓴 섹션이면
    'start_lot_id'를 채워서 대시보드에서 LOT 선택 없이 바로 시작할 수 있게 한다."""
    rows = db.execute("""
        SELECT ls.*, l.model_name, l.lot_number
        FROM production_lot_sections ls
        JOIN production_lots l ON ls.lot_id = l.id
        WHERE ls.status IN ('진행중','일시정지')
    """).fetchall()
    by_section = {r['section']: r for r in rows}

    active_lots = db.execute(
        "SELECT id, model_name, lot_number FROM production_lots WHERE status='진행중' ORDER BY created_at DESC"
    ).fetchall()
    used_by_sole_active_lot = set()
    if len(active_lots) == 1:
        used_by_sole_active_lot = {r['section'] for r in db.execute(
            "SELECT section FROM production_lot_sections WHERE lot_id=?", (active_lots[0]['id'],)
        ).fetchall()}

    result = []
    for section in SECTIONS:
        ls = by_section.get(section)
        if ls:
            stats = compute_section_stats(db, ls['id'], lot_section_row=ls)
            result.append({
                'section': section,
                'running': ls['status'] == '진행중',
                'paused': ls['status'] == '일시정지',
                'lot_id': ls['lot_id'],
                'model_name': ls['model_name'],
                'lot_number': ls['lot_number'],
                'uptime_rate': stats['uptime_rate'],
                'fault_count': stats['fault_count'],
            })
        else:
            start_lot_id = None
            if len(active_lots) == 1 and section not in used_by_sole_active_lot:
                start_lot_id = active_lots[0]['id']
            result.append({
                'section': section, 'running': False, 'paused': False,
                'start_lot_id': start_lot_id,
                'multiple_active_lots': len(active_lots) > 1,
            })
    return result


def _subtract_intervals(base_list, cutters):
    """base_list의 각 구간에서 cutters와 겹치는 부분을 잘라내고 남은 조각을 반환."""
    out = []
    for b in base_list:
        pieces = [(b['start'], b['end'])]
        for c in cutters:
            next_pieces = []
            for ps, pe in pieces:
                if c['end'] <= ps or c['start'] >= pe:
                    next_pieces.append((ps, pe))
                    continue
                if c['start'] > ps:
                    next_pieces.append((ps, c['start']))
                if c['end'] < pe:
                    next_pieces.append((c['end'], pe))
            pieces = next_pieces
        out += [{'type': b['type'], 'start': ps, 'end': pe, 'label': b['label']} for ps, pe in pieces if pe > ps]
    return out


def compute_section_timeline(db, ls):
    """섹션 가동구간을 시간순으로 정상가동/휴식시간/가동정지(다운타임)/고장정지 구간으로 나눈 목록을 반환.
    우선순위는 고장 > 가동정지(다운타임) > 휴식시간 순 (실제로 멈춘 사실이 더 중요하므로)."""
    win_start = _parse_dt(ls['started_at'])
    win_end = _parse_dt(ls['ended_at']) if ls['ended_at'] else datetime.now()
    if win_end <= win_start:
        return []

    break_intervals = [
        {'type': 'break', 'start': iv['start'], 'end': iv['end'], 'label': iv['label']}
        for iv in _excluded_intervals(db, ls['id'], win_start, win_end)
    ]

    downtime_intervals = [
        {'type': 'downtime', 'start': p['start'], 'end': p['end'], 'label': p['label']}
        for p in _downtime_pause_intervals(db, ls['id'], win_start, win_end)
    ]

    faults = db.execute("""
        SELECT occurred_at, completed_at, symptom, cause FROM fault_history
        WHERE lot_section_id = ? ORDER BY occurred_at
    """, (ls['id'],)).fetchall()
    fault_intervals = []
    for f in faults:
        s = max(win_start, _parse_dt(f['occurred_at']))
        e = min(win_end, _parse_dt(f['completed_at']) if f['completed_at'] else win_end)
        if e > s:
            label = (f['symptom'] or f['cause'] or '고장').split(',')[0].strip()
            fault_intervals.append({'type': 'fault', 'start': s, 'end': e, 'label': label})

    # 우선순위: 고장 > 가동정지(다운타임) > 휴식시간 - 겹치는 조각은 우선순위 낮은 쪽에서 잘라낸다
    downtime_intervals = _subtract_intervals(downtime_intervals, fault_intervals)
    break_intervals = _subtract_intervals(break_intervals, fault_intervals + downtime_intervals)

    all_intervals = sorted(break_intervals + downtime_intervals + fault_intervals, key=lambda x: x['start'])

    segments = []
    cursor = win_start
    for iv in all_intervals:
        if iv['start'] > cursor:
            segments.append({'type': 'normal', 'start': cursor, 'end': iv['start'], 'label': '정상가동'})
        segments.append(iv)
        cursor = max(cursor, iv['end'])
    if cursor < win_end:
        segments.append({'type': 'normal', 'start': cursor, 'end': win_end, 'label': '정상가동'})

    return segments


def log_action(action, target_type, target_id, detail=None):
    db = get_db()
    db.execute(
        "INSERT INTO audit_log (actor, action, target_type, target_id, detail) VALUES (?,?,?,?,?)",
        (session.get('actor_name', '이름 미선택'), action, target_type, target_id, detail)
    )
    db.commit()
    db.close()


def get_local_ip():
    # Tailscale 등으로 암호화 접속을 강제하려면 QR_SERVER_HOST 환경변수(IP 또는
    # MagicDNS 호스트명)를 설정한다. 실행.bat 참고. 설정 없으면 기존처럼 LAN IP 자동 감지.
    override = os.environ.get('QR_SERVER_HOST', '').strip()
    if override:
        return override
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'


def parse_occurred_at(raw):
    """datetime-local 입력값('YYYY-MM-DDTHH:MM')을 DB 저장 형식으로 변환. 값이 없으면 현재 시각."""
    raw = (raw or '').strip()
    if not raw:
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    raw = raw.replace('T', ' ')
    if len(raw) == 16:
        raw += ':00'
    return raw


def parse_optional_datetime(raw):
    """datetime-local 입력값을 DB 저장 형식으로 변환. 값이 없으면 None (완료 안 됨)."""
    raw = (raw or '').strip()
    if not raw:
        return None
    raw = raw.replace('T', ' ')
    if len(raw) == 16:
        raw += ':00'
    return raw


def generate_qr(eq_id, host_url=None):
    # 항상 LAN IP 사용 — 모바일(휴대폰)에서도 스캔 가능하도록
    ip = get_local_ip()
    url = f"http://{ip}:5001/fault/register/{eq_id}"
    qr = qrcode.QRCode(version=1, box_size=10, border=4,
                        error_correction=qrcode.constants.ERROR_CORRECT_H)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    path = os.path.join(QR_DIR, f'eq_{eq_id}.png')
    img.save(path)
    # 이 함수는 최초 생성(파일 없음) 또는 관리자의 명시적 "QR 재생성" 클릭 때만 호출되므로,
    # 호출될 때마다 "실제로 QR이 (재)생성된 시각"으로 기록해도 된다 (조회만으로는 갱신되지 않음).
    db = get_db()
    db.execute("UPDATE equipment SET qr_reset_at=datetime('now','localtime'), qr_last_ip=? WHERE id=?", (ip, eq_id))
    db.commit()
    db.close()
    return path


# ──────────────────────────────────────────────
# 대시보드
# ──────────────────────────────────────────────
@app.route('/')
def index():
    db = get_db()

    recent_faults = db.execute("""
        SELECT id, eq_number, eq_name, occurred_at, symptom, worker, grade, status
        FROM fault_history ORDER BY occurred_at DESC LIMIT 8
    """).fetchall()
    recent_production = db.execute("""
        SELECT actor, action, detail, created_at, target_id
        FROM audit_log
        WHERE target_type='production_lot' AND action IN
            ('LOT 등록','섹션 가동시작','섹션 가동정지','섹션 가동재개','섹션 가동종료','LOT 종료')
        ORDER BY created_at DESC LIMIT 8
    """).fetchall()

    total     = db.execute("SELECT COUNT(*) FROM fault_history").fetchone()[0]
    this_week = db.execute("""
        SELECT COUNT(*) FROM fault_history
        WHERE strftime('%Y-%W', occurred_at) = strftime('%Y-%W', 'now','localtime')
    """).fetchone()[0]
    this_month = db.execute("""
        SELECT COUNT(*) FROM fault_history
        WHERE strftime('%Y-%m', occurred_at) = strftime('%Y-%m', 'now','localtime')
    """).fetchone()[0]

    top_eq = db.execute("""
        SELECT eq_number, eq_name, COUNT(*) as cnt
        FROM fault_history GROUP BY equipment_id
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    top_symptom = db.execute("""
        SELECT symptom, COUNT(*) as cnt
        FROM fault_history GROUP BY symptom
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    top_worker = db.execute("""
        SELECT worker, COUNT(*) as cnt
        FROM fault_history WHERE worker IS NOT NULL AND worker != ''
        GROUP BY worker ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    top_parts = db.execute("""
        SELECT part_name, SUM(quantity) as cnt
        FROM used_parts GROUP BY part_name
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    completed_parts = db.execute("""
        SELECT up.part_name, SUM(up.quantity) as cnt, COUNT(DISTINCT up.fault_id) as cases
        FROM used_parts up
        JOIN fault_history f ON up.fault_id = f.id
        WHERE f.status = '완료'
        GROUP BY up.part_name ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    grade_cnt = db.execute("""
        SELECT grade, COUNT(*) as cnt
        FROM fault_history GROUP BY grade ORDER BY grade
    """).fetchall()

    weekly = db.execute("""
        SELECT strftime('%Y-W%W', occurred_at) as week, COUNT(*) as cnt
        FROM fault_history
        WHERE occurred_at >= date('now','localtime','-12 weeks')
        GROUP BY week ORDER BY week
    """).fetchall()

    monthly = db.execute("""
        SELECT strftime('%Y-%m', occurred_at) as month, COUNT(*) as cnt
        FROM fault_history
        WHERE occurred_at >= date('now','localtime','-12 months')
        GROUP BY month ORDER BY month
    """).fetchall()

    available_months = [
        {'value': mstr, 'label': f"{mstr.split('-')[0]}년 {int(mstr.split('-')[1])}월"}
        for mstr in last_n_months(12)
    ]

    selected_month = request.args.get('month', '').strip() or date.today().strftime('%Y-%m')
    sy, sm = selected_month.split('-')
    selected_month_label = f"{sy}년 {int(sm)}월"

    _, last_day = calendar.monthrange(int(sy), int(sm))
    day_rows = db.execute("""
        SELECT CAST(strftime('%d', occurred_at) AS INTEGER) as d, COUNT(*) as cnt
        FROM fault_history
        WHERE strftime('%Y-%m', occurred_at) = ?
        GROUP BY d
    """, (selected_month,)).fetchall()
    day_counts = {r['d']: r['cnt'] for r in day_rows}
    daily_trend = [{'day': f'{d}일', 'cnt': day_counts.get(d, 0)} for d in range(1, last_day + 1)]

    weekly_trend = db.execute("""
        SELECT strftime('%Y-W%W', occurred_at) as week, COUNT(*) as cnt
        FROM fault_history
        WHERE strftime('%Y-%m', occurred_at) = ?
        GROUP BY week ORDER BY week
    """, (selected_month,)).fetchall()

    eq_chart = db.execute("""
        SELECT eq_number || ' ' || eq_name as label, COUNT(*) as cnt
        FROM fault_history GROUP BY equipment_id
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    symptom_chart = db.execute("""
        SELECT symptom as label, COUNT(*) as cnt
        FROM fault_history GROUP BY symptom
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    parts_chart = db.execute("""
        SELECT part_name as label, SUM(quantity) as cnt
        FROM used_parts GROUP BY part_name
        ORDER BY cnt DESC LIMIT 10
    """).fetchall()

    status_cnt = db.execute("""
        SELECT status, COUNT(*) as cnt FROM fault_history
        GROUP BY status
    """).fetchall()

    section_status = get_section_dashboard_status(db)

    db.close()

    today_dt = date.today()
    week_start = (today_dt - timedelta(days=today_dt.weekday())).isoformat()
    month_start = today_dt.replace(day=1).isoformat()

    return render_template('index.html',
        total=total, this_week=this_week, this_month=this_month,
        top_eq=top_eq, top_symptom=top_symptom,
        top_worker=top_worker, top_parts=top_parts,
        completed_parts=completed_parts,
        grade_cnt=grade_cnt, status_cnt=status_cnt,
        weekly=weekly, monthly=monthly,
        available_months=available_months,
        selected_month=selected_month, selected_month_label=selected_month_label,
        daily_trend=daily_trend, weekly_trend=weekly_trend,
        eq_chart=eq_chart, symptom_chart=symptom_chart,
        parts_chart=parts_chart,
        week_start=week_start, month_start=month_start,
        section_status=section_status, pause_reasons=PAUSE_REASONS,
        kakao_connected=_kakao_connected(), kakao_accounts=_kakao_load().get('accounts', []),
        notify_types=NOTIFY_TYPES, notify_settings=_load_notify_settings(),
        recent_faults=recent_faults, recent_production=recent_production)


@app.route('/dashboard/sections.json')
def dashboard_sections_json():
    db = get_db()
    section_status = get_section_dashboard_status(db)
    db.close()
    return jsonify(ok=True, sections=section_status)


# ──────────────────────────────────────────────
# 통합검색 (설비 + 고장이력)
# ──────────────────────────────────────────────
@app.route('/search')
def global_search():
    q = request.args.get('q', '').strip()
    equipments, faults = [], []
    if q:
        db = get_db()
        like = f'%{q}%'
        equipments = db.execute("""
            SELECT * FROM equipment
            WHERE eq_number LIKE ? OR eq_name LIKE ? OR location LIKE ? OR section LIKE ?
            ORDER BY section, eq_number
        """, [like, like, like, like]).fetchall()
        faults = db.execute("""
            SELECT * FROM fault_history
            WHERE eq_number LIKE ? OR eq_name LIKE ? OR symptom LIKE ? OR worker LIKE ? OR cause LIKE ?
            ORDER BY occurred_at DESC LIMIT 50
        """, [like, like, like, like, like]).fetchall()
        db.close()
    return render_template('search.html', q=q, equipments=equipments, faults=faults)


# ──────────────────────────────────────────────
# 설비 관리
# ──────────────────────────────────────────────
@app.route('/equipment/list')
def equipment_list():
    q = request.args.get('q', '').strip()
    sec = request.args.get('section', '').strip()
    db = get_db()
    sql = "SELECT * FROM equipment WHERE 1=1"
    params = []
    if q:
        sql += " AND (eq_number LIKE ? OR eq_name LIKE ? OR location LIKE ?)"
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if sec:
        sql += " AND section = ?"
        params.append(sec)
    sql += " ORDER BY section, eq_number"
    equipments = db.execute(sql, params).fetchall()
    db.close()
    return render_template('equipment/list.html',
                           equipments=equipments, sections=SECTIONS, q=q, sec=sec)


@app.route('/equipment/add', methods=['GET', 'POST'])
def equipment_add():
    if not session.get('edit_authorized'):
        flash('설비 등록은 관리자 인증이 필요합니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(url_for('equipment_list'))
    if request.method == 'POST':
        eq_number = request.form.get('eq_number', '').strip()
        eq_name   = request.form.get('eq_name', '').strip()
        section   = request.form.get('section', '').strip()
        location  = request.form.get('location', '').strip()
        note      = request.form.get('note', '').strip()

        if not eq_number or not eq_name or not section:
            flash('설비번호, 설비명, 섹션은 필수입니다.', 'danger')
            return render_template('equipment/add.html', sections=SECTIONS)

        db = get_db()
        cur = db.execute(
            "INSERT INTO equipment (eq_number,eq_name,section,location,note) VALUES (?,?,?,?,?)",
            (eq_number, eq_name, section, location, note)
        )
        eq_id = cur.lastrowid
        db.commit()
        db.close()
        log_action('등록', '설비', eq_id, f'{eq_number} {eq_name}')

        host_url = request.host_url
        generate_qr(eq_id, host_url)

        flash(f'설비 "{eq_name}" 등록 완료. QR코드 생성됨.', 'success')
        return redirect(url_for('equipment_list'))

    return render_template('equipment/add.html', sections=SECTIONS)


@app.route('/equipment/edit/<int:eq_id>', methods=['GET', 'POST'])
def equipment_edit(eq_id):
    if not session.get('edit_authorized'):
        flash('설비 수정은 관리자 인증이 필요합니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(url_for('equipment_detail', eq_id=eq_id))
    db = get_db()
    eq = db.execute("SELECT * FROM equipment WHERE id=?", (eq_id,)).fetchone()
    if not eq:
        db.close()
        abort(404)

    if request.method == 'POST':
        eq_number = request.form.get('eq_number', '').strip()
        eq_name   = request.form.get('eq_name', '').strip()
        section   = request.form.get('section', '').strip()
        location  = request.form.get('location', '').strip()
        note      = request.form.get('note', '').strip()

        if not eq_number or not eq_name or not section:
            flash('설비번호, 설비명, 섹션은 필수입니다.', 'danger')
            db.close()
            return render_template('equipment/edit.html', eq=eq, sections=SECTIONS)

        db.execute(
            "UPDATE equipment SET eq_number=?,eq_name=?,section=?,location=?,note=? WHERE id=?",
            (eq_number, eq_name, section, location, note, eq_id)
        )
        db.commit()
        db.close()
        log_action('수정', '설비', eq_id, f'{eq_number} {eq_name}')

        host_url = request.host_url
        generate_qr(eq_id, host_url)

        flash(f'설비 정보가 수정되었습니다.', 'success')
        return redirect(url_for('equipment_list'))

    db.close()
    return render_template('equipment/edit.html', eq=eq, sections=SECTIONS)


@app.route('/equipment/delete/<int:eq_id>', methods=['POST'])
def equipment_delete(eq_id):
    if not session.get('edit_authorized'):
        flash('삭제 권한이 없습니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(request.referrer or url_for('equipment_list'))
    db = get_db()
    eq = db.execute("SELECT * FROM equipment WHERE id=?", (eq_id,)).fetchone()
    if not eq:
        db.close()
        abort(404)
    fault_count = db.execute(
        "SELECT COUNT(*) FROM fault_history WHERE equipment_id=?", (eq_id,)
    ).fetchone()[0]
    if fault_count > 0:
        db.close()
        flash(f'이 설비에는 고장이력이 {fault_count}건 있어 삭제할 수 없습니다. '
              f'먼저 고장이력을 삭제한 후 다시 시도해주세요.', 'danger')
        return redirect(url_for('equipment_detail', eq_id=eq_id))
    db.execute("DELETE FROM equipment WHERE id=?", (eq_id,))
    db.commit()
    db.close()
    log_action('삭제', '설비', eq_id, f"{eq['eq_number']} {eq['eq_name']}")
    qr_path = os.path.join(QR_DIR, f'eq_{eq_id}.png')
    if os.path.exists(qr_path):
        os.remove(qr_path)
    flash('설비가 삭제되었습니다.', 'warning')
    return redirect(url_for('equipment_list'))


@app.route('/equipment/<int:eq_id>')
def equipment_detail(eq_id):
    db = get_db()
    eq = db.execute("SELECT * FROM equipment WHERE id=?", (eq_id,)).fetchone()
    if not eq:
        db.close()
        abort(404)
    recent_faults = db.execute(
        "SELECT * FROM fault_history WHERE equipment_id=? ORDER BY occurred_at DESC LIMIT 5",
        (eq_id,)
    ).fetchall()
    fault_count = db.execute(
        "SELECT COUNT(*) FROM fault_history WHERE equipment_id=?", (eq_id,)
    ).fetchone()[0]
    db.close()
    return render_template('equipment/detail.html',
                           eq=eq, recent_faults=recent_faults, fault_count=fault_count)


# ──────────────────────────────────────────────
# QR 생성 / 출력
# ──────────────────────────────────────────────
@app.route('/qr/generate/<int:eq_id>')
def qr_generate(eq_id):
    db = get_db()
    eq = db.execute("SELECT * FROM equipment WHERE id=?", (eq_id,)).fetchone()
    db.close()
    if not eq:
        abort(404)
    host_url = request.host_url
    generate_qr(eq_id, host_url)
    flash('QR코드가 재생성되었습니다.', 'success')
    return redirect(url_for('qr_print', eq_id=eq_id))


@app.route('/qr/print/<int:eq_id>')
def qr_print(eq_id):
    db = get_db()
    eq = db.execute("SELECT * FROM equipment WHERE id=?", (eq_id,)).fetchone()
    db.close()
    if not eq:
        abort(404)
    qr_path = os.path.join(QR_DIR, f'eq_{eq_id}.png')
    if not os.path.exists(qr_path):
        generate_qr(eq_id, request.host_url)
    return render_template('qr/print.html', eq=eq)


@app.route('/qr/image/<int:eq_id>')
def qr_image(eq_id):
    path = os.path.join(QR_DIR, f'eq_{eq_id}.png')
    if not os.path.exists(path):
        db = get_db()
        eq = db.execute("SELECT * FROM equipment WHERE id=?", (eq_id,)).fetchone()
        db.close()
        if not eq:
            abort(404)
        generate_qr(eq_id, request.host_url)
    return send_file(path, mimetype='image/png')


@app.route('/qr/all')
def qr_all():
    db = get_db()
    sec = request.args.get('section', '').strip()
    sql = "SELECT * FROM equipment"
    params = []
    if sec:
        sql += " WHERE section=?"
        params.append(sec)
    sql += " ORDER BY section, eq_number"
    equipments = db.execute(sql, params).fetchall()
    db.close()
    for eq in equipments:
        qr_path = os.path.join(QR_DIR, f'eq_{eq["id"]}.png')
        if not os.path.exists(qr_path):
            generate_qr(eq['id'], request.host_url)
    return render_template('qr/print_all.html', equipments=equipments, sections=SECTIONS, sec=sec)


# ──────────────────────────────────────────────
# 고장 등록 — 설비 선택 (대시보드 직접 접근용)
# ──────────────────────────────────────────────
@app.route('/fault/select')
def fault_select():
    q   = request.args.get('q', '').strip()
    sec = request.args.get('section', '').strip()
    db  = get_db()
    sql = "SELECT * FROM equipment WHERE 1=1"
    params = []
    if q:
        sql += " AND (eq_number LIKE ? OR eq_name LIKE ? OR location LIKE ?)"
        params += [f'%{q}%', f'%{q}%', f'%{q}%']
    if sec:
        sql += " AND section = ?"
        params.append(sec)
    sql += " ORDER BY section, eq_number"
    equipments = db.execute(sql, params).fetchall()
    db.close()
    return render_template('fault/select_equipment.html',
                           equipments=equipments, sections=SECTIONS, q=q, sec=sec)


# ──────────────────────────────────────────────
# 고장 등록
# ──────────────────────────────────────────────
@app.route('/fault/register/<int:eq_id>', methods=['GET', 'POST'])
def fault_register(eq_id):
    db = get_db()
    eq = db.execute("SELECT * FROM equipment WHERE id=?", (eq_id,)).fetchone()
    if not eq:
        db.close()
        abort(404)

    if request.method == 'POST':
        symptom_list = request.form.getlist('symptom[]')
        symptom      = ', '.join([s for s in symptom_list if s])
        cause        = request.form.get('cause', '').strip()
        action_detail= request.form.get('action_detail', '').strip()
        worker_sel   = request.form.get('worker_select', '').strip()
        worker_custom= request.form.get('worker_custom', '').strip()
        worker       = worker_custom if worker_sel == '__new__' else worker_sel
        completed_at = parse_optional_datetime(request.form.get('completed_at', ''))
        grade        = request.form.get('grade', '').strip()
        grade_note   = request.form.get('grade_note', '').strip() if grade == 'D' else None
        status       = request.form.get('status', '미조치').strip()
        status_note  = request.form.get('status_note', '').strip() if status == '대기' else None
        occurred_at  = parse_occurred_at(request.form.get('occurred_at', ''))

        if worker:
            register_worker(db, worker)

        active_lot_section = get_active_lot_section(db, eq['section'])
        lot_section_id = active_lot_section['id'] if active_lot_section else None
        lot_id = active_lot_section['lot_id'] if active_lot_section else None

        cur = db.execute("""
            INSERT INTO fault_history
            (equipment_id,eq_number,eq_name,occurred_at,symptom,cause,action_detail,worker,completed_at,grade,grade_note,photo_fault,photo_action,status,status_note,lot_id,lot_section_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (eq_id, eq['eq_number'], eq['eq_name'], occurred_at,
              symptom, cause, action_detail, worker, completed_at, grade, grade_note,
              None, None, status, status_note, lot_id, lot_section_id))
        fault_id = cur.lastrowid

        # 고장 사진 저장 (최대 5장)
        photo_save_failed = False
        def save_photos(files_key, labels_key, ptype):
            nonlocal photo_save_failed
            files  = request.files.getlist(files_key)
            labels = request.form.getlist(labels_key)
            for i, pf in enumerate(files[:5]):
                if pf and pf.filename and allowed_file(pf.filename):
                    label = (labels[i] if i < len(labels) else '').strip()
                    ts = datetime.now().strftime('%Y%m%d%H%M%S%f')[:17]
                    base = os.path.splitext(secure_filename(pf.filename))[0]
                    fn = f"p{eq_id}_{ptype}_{ts}_{i}_{base}.jpg"
                    try:
                        save_and_compress_image(pf, os.path.join(PHOTOS_DIR, fn))
                        db.execute(
                            "INSERT INTO fault_photos (fault_id,filename,label,photo_type) VALUES (?,?,?,?)",
                            (fault_id, fn, label, ptype)
                        )
                    except Exception:
                        logging.exception('photo save failed: %s', fn)
                        photo_save_failed = True

        save_photos('fault_photos[]', 'fault_photo_labels[]', 'fault')
        save_photos('action_photos[]', 'action_photo_labels[]', 'action')

        part_names = request.form.getlist('part_name[]')
        part_qtys  = request.form.getlist('part_qty[]')
        part_notes = request.form.getlist('part_note[]')
        for pname, pqty, pnote in zip(part_names, part_qtys, part_notes):
            pname = pname.strip()
            if pname:
                try:
                    qty = int(pqty) if pqty else 1
                except ValueError:
                    qty = 1
                db.execute(
                    "INSERT INTO used_parts (fault_id,part_name,quantity,note) VALUES (?,?,?,?)",
                    (fault_id, pname, qty, pnote.strip())
                )

        db.commit()
        db.close()
        notify_fault_by_grade(grade, fault_id, eq['eq_number'], eq['eq_name'], symptom, status, worker, action_detail)
        flash('고장이 등록되었습니다.', 'success')
        if photo_save_failed:
            flash('일부 사진 저장에 실패했습니다. 필요하면 수정 화면에서 다시 첨부해주세요.', 'warning')
        return redirect(url_for('equipment_detail', eq_id=eq_id))

    workers = get_workers(db)
    active_lot_section = db.execute("""
        SELECT ls.*, l.model_name, l.lot_number
        FROM production_lot_sections ls
        JOIN production_lots l ON ls.lot_id = l.id
        WHERE ls.section=? AND ls.status='진행중'
        ORDER BY ls.started_at DESC LIMIT 1
    """, (eq['section'],)).fetchone()
    paused_lot_section = None
    if not active_lot_section:
        paused_lot_section = db.execute("""
            SELECT ls.*, l.model_name, l.lot_number
            FROM production_lot_sections ls
            JOIN production_lots l ON ls.lot_id = l.id
            WHERE ls.section=? AND ls.status='일시정지'
            ORDER BY ls.started_at DESC LIMIT 1
        """, (eq['section'],)).fetchone()
    db.close()
    return render_template('fault/register.html',
                           eq=eq, symptoms=SYMPTOMS, grades=GRADES,
                           parts=PARTS, statuses=STATUSES, workers=workers,
                           today=date.today().isoformat(),
                           now_dt=datetime.now().strftime('%Y-%m-%dT%H:%M'),
                           active_lot_section=active_lot_section,
                           paused_lot_section=paused_lot_section)


# ──────────────────────────────────────────────
# 고장 이력 조회
# ──────────────────────────────────────────────
@app.route('/fault/list')
def fault_list():
    q         = request.args.get('q', '').strip()
    sym       = request.args.get('symptom', '').strip()
    worker    = request.args.get('worker', '').strip()
    grade     = request.args.get('grade', '').strip()
    status    = request.args.get('status', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    eq_id     = request.args.get('eq_id', '').strip()

    sql = "SELECT * FROM fault_history WHERE 1=1"
    params = []

    if q:
        sql += " AND (eq_number LIKE ? OR eq_name LIKE ? OR symptom LIKE ? OR worker LIKE ?)"
        params += [f'%{q}%'] * 4
    if sym:
        sql += " AND symptom LIKE ?"
        params.append(f'%{sym}%')
    if worker:
        sql += " AND worker LIKE ?"
        params.append(f'%{worker}%')
    if grade:
        sql += " AND grade=?"
        params.append(grade)
    if status:
        sql += " AND (status=? OR (status IS NULL AND ?='미조치'))"
        params += [status, status]
    if date_from:
        sql += " AND date(occurred_at) >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND date(occurred_at) <= ?"
        params.append(date_to)
    if eq_id:
        sql += " AND equipment_id=?"
        params.append(eq_id)

    PAGE_SIZE = 20
    try:
        page = max(1, int(request.args.get('page', 1)))
    except ValueError:
        page = 1

    db = get_db()
    total_count = db.execute(f"SELECT COUNT(*) c FROM ({sql})", params).fetchone()['c']
    total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)
    page = min(page, total_pages)

    sql += " ORDER BY occurred_at DESC LIMIT ? OFFSET ?"
    params_paged = params + [PAGE_SIZE, (page - 1) * PAGE_SIZE]
    faults = db.execute(sql, params_paged).fetchall()
    workers = db.execute(
        "SELECT DISTINCT worker FROM fault_history WHERE worker IS NOT NULL AND worker != '' ORDER BY worker"
    ).fetchall()
    db.close()

    def page_url(p):
        args = request.args.to_dict()
        args['page'] = p
        return '?' + urlencode(args)

    return render_template('fault/list.html',
                           faults=faults, symptoms=SYMPTOMS, workers=workers,
                           grades=GRADES, statuses=STATUSES,
                           q=q, sym=sym, worker=worker, grade=grade,
                           status=status, date_from=date_from, date_to=date_to,
                           page=page, total_pages=total_pages, total_count=total_count,
                           page_url=page_url)


@app.route('/fault/<int:fault_id>')
def fault_detail(fault_id):
    db = get_db()
    fault = db.execute("SELECT * FROM fault_history WHERE id=?", (fault_id,)).fetchone()
    if not fault:
        db.close()
        abort(404)
    parts         = db.execute("SELECT * FROM used_parts WHERE fault_id=?", (fault_id,)).fetchall()
    fault_photos  = db.execute(
        "SELECT * FROM fault_photos WHERE fault_id=? AND photo_type='fault' ORDER BY id", (fault_id,)).fetchall()
    action_photos = db.execute(
        "SELECT * FROM fault_photos WHERE fault_id=? AND photo_type='action' ORDER BY id", (fault_id,)).fetchall()
    db.close()
    return render_template('fault/detail.html', fault=fault, parts=parts,
                           fault_photos=fault_photos, action_photos=action_photos)


# ──────────────────────────────────────────────
# 수정 권한 인증
# ──────────────────────────────────────────────
@app.route('/fault/auth', methods=['POST'])
def fault_auth():
    ip = request.remote_addr
    remaining = _login_lock_remaining(ip)
    if remaining > 0:
        flash(f'비밀번호를 너무 많이 틀렸습니다. {int(remaining // 60) + 1}분 후 다시 시도하세요.', 'danger')
        return redirect(request.referrer or url_for('fault_list'))

    pw           = request.form.get('password', '').strip()
    actor_sel    = request.form.get('actor', '').strip()
    actor_custom = request.form.get('actor_custom', '').strip()
    actor        = actor_custom if actor_sel == '__new__' else actor_sel
    next_url = request.form.get('next', '/')
    if check_admin_password(pw):
        _register_login_success(ip)
        if actor:
            db = get_db()
            register_worker(db, actor)
            db.commit()
            db.close()
        session.permanent = True
        session['edit_authorized'] = True
        session['actor_name'] = actor or '이름 미선택'
        return redirect(next_url)
    _register_login_failure(ip)
    flash('비밀번호가 올바르지 않습니다.', 'danger')
    return redirect(request.referrer or url_for('fault_list'))


@app.route('/worker/delete', methods=['POST'])
def worker_delete():
    ip = request.remote_addr
    remaining = _login_lock_remaining(ip)
    if remaining > 0:
        return jsonify(ok=False, error=f'비밀번호를 너무 많이 틀렸습니다. {int(remaining // 60) + 1}분 후 다시 시도하세요.'), 429

    pw   = request.form.get('password', '').strip()
    name = request.form.get('name', '').strip()
    if not check_admin_password(pw):
        _register_login_failure(ip)
        return jsonify(ok=False, error='비밀번호가 올바르지 않습니다.'), 403
    _register_login_success(ip)
    if not name:
        return jsonify(ok=False, error='삭제할 이름을 선택하세요.'), 400
    db = get_db()
    db.execute("DELETE FROM workers WHERE name=?", (name,))
    db.commit()
    db.close()
    return jsonify(ok=True)


@app.route('/login', methods=['GET', 'POST'])
def login():
    next_url = request.values.get('next') or url_for('index')
    if request.method == 'POST':
        ip = request.remote_addr
        remaining = _login_lock_remaining(ip)
        if remaining > 0:
            flash(f'비밀번호를 너무 많이 틀렸습니다. {int(remaining // 60) + 1}분 후 다시 시도하세요.', 'danger')
            return redirect(url_for('login', next=next_url))

        pw           = request.form.get('password', '').strip()
        actor_sel    = request.form.get('actor', '').strip()
        actor_custom = request.form.get('actor_custom', '').strip()
        actor        = actor_custom if actor_sel == '__new__' else actor_sel
        if check_admin_password(pw):
            _register_login_success(ip)
            if actor:
                db = get_db()
                register_worker(db, actor)
                db.commit()
                db.close()
            session.permanent = True
            session['edit_authorized'] = True
            session['actor_name'] = actor or '이름 미선택'
            return redirect(next_url)
        _register_login_failure(ip)
        flash('비밀번호가 올바르지 않습니다.', 'danger')
        return redirect(url_for('login', next=next_url))
    db = get_db()
    workers = get_workers(db)
    db.close()
    return render_template('login.html', workers=workers, next=next_url)


@app.route('/settings/password', methods=['POST'])
def change_admin_password():
    if not session.get('edit_authorized'):
        flash('로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login'))
    current_pw = request.form.get('current_password', '').strip()
    new_pw     = request.form.get('new_password', '').strip()
    new_pw2    = request.form.get('new_password_confirm', '').strip()
    next_url   = request.form.get('next') or url_for('index')

    if not check_admin_password(current_pw):
        flash('현재 비밀번호가 올바르지 않습니다.', 'danger')
        return redirect(next_url)
    if len(new_pw) < 4:
        flash('새 비밀번호는 4자 이상이어야 합니다.', 'danger')
        return redirect(next_url)
    if new_pw != new_pw2:
        flash('새 비밀번호가 서로 일치하지 않습니다.', 'danger')
        return redirect(next_url)

    set_admin_password(new_pw)
    flash('비밀번호가 변경되었습니다.', 'success')
    return redirect(next_url)


@app.route('/logout')
def logout():
    session.pop('edit_authorized', None)
    session.pop('actor_name', None)
    flash('로그아웃되었습니다.', 'success')
    return redirect(url_for('login'))


# ──────────────────────────────────────────────
# 고장 이력 수정
# ──────────────────────────────────────────────
@app.route('/fault/edit/<int:fault_id>', methods=['GET', 'POST'])
def fault_edit(fault_id):
    if not session.get('edit_authorized'):
        flash('수정 권한이 없습니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(url_for('fault_detail', fault_id=fault_id))

    db = get_db()
    fault = db.execute("SELECT * FROM fault_history WHERE id=?", (fault_id,)).fetchone()
    if not fault:
        db.close()
        abort(404)

    if request.method == 'POST':
        symptom_list = request.form.getlist('symptom[]')
        symptom      = ', '.join([s for s in symptom_list if s])
        cause        = request.form.get('cause', '').strip()
        action_detail= request.form.get('action_detail', '').strip()
        worker_sel   = request.form.get('worker_select', '').strip()
        worker_custom= request.form.get('worker_custom', '').strip()
        worker       = worker_custom if worker_sel == '__new__' else worker_sel
        completed_at = parse_optional_datetime(request.form.get('completed_at', ''))
        grade        = request.form.get('grade', '').strip()
        grade_note   = request.form.get('grade_note', '').strip() if grade == 'D' else None
        status       = request.form.get('status', '미조치').strip()
        status_note  = request.form.get('status_note', '').strip() if status == '대기' else None
        occurred_at  = parse_occurred_at(request.form.get('occurred_at', ''))

        if worker:
            register_worker(db, worker)

        # 기존 사진 삭제 요청 처리
        delete_ids = request.form.getlist('delete_photo[]')
        for pid in delete_ids:
            p = db.execute("SELECT filename FROM fault_photos WHERE id=? AND fault_id=?",
                           (pid, fault_id)).fetchone()
            if p:
                fp = os.path.join(PHOTOS_DIR, p['filename'])
                if os.path.exists(fp):
                    os.remove(fp)
                db.execute("DELETE FROM fault_photos WHERE id=?", (pid,))

        db.execute("""
            UPDATE fault_history SET
              occurred_at=?, symptom=?, cause=?, action_detail=?, worker=?,
              completed_at=?, grade=?, grade_note=?, status=?, status_note=?
            WHERE id=?
        """, (occurred_at, symptom, cause, action_detail, worker, completed_at,
              grade, grade_note, status, status_note, fault_id))

        # 새 사진 추가 (고장/완료 각 타입별 최대 5장)
        eq_id_edit = fault['equipment_id']
        photo_save_failed = False
        def save_edit_photos(files_key, labels_key, ptype):
            nonlocal photo_save_failed
            files  = request.files.getlist(files_key)
            labels = request.form.getlist(labels_key)
            exist = db.execute(
                "SELECT COUNT(*) FROM fault_photos WHERE fault_id=? AND photo_type=?",
                (fault_id, ptype)).fetchone()[0]
            for i, pf in enumerate(files):
                if exist >= 5:
                    break
                if pf and pf.filename and allowed_file(pf.filename):
                    label = (labels[i] if i < len(labels) else '').strip()
                    ts = datetime.now().strftime('%Y%m%d%H%M%S%f')[:17]
                    base = os.path.splitext(secure_filename(pf.filename))[0]
                    fn = f"p{eq_id_edit}_{ptype}_{ts}_{i}_{base}.jpg"
                    try:
                        save_and_compress_image(pf, os.path.join(PHOTOS_DIR, fn))
                        db.execute(
                            "INSERT INTO fault_photos (fault_id,filename,label,photo_type) VALUES (?,?,?,?)",
                            (fault_id, fn, label, ptype))
                        exist += 1
                    except Exception:
                        logging.exception('photo save failed: %s', fn)
                        photo_save_failed = True

        save_edit_photos('fault_photos[]', 'fault_photo_labels[]', 'fault')
        save_edit_photos('action_photos[]', 'action_photo_labels[]', 'action')
        db.commit()
        db.close()
        if grade != fault['grade']:
            notify_fault_by_grade(grade, fault_id, fault['eq_number'], fault['eq_name'], symptom, status, worker, action_detail)
        log_action('수정', '고장이력', fault_id,
                   f"{fault['eq_number']} {fault['eq_name']} / 등급 {fault['grade'] or '-'}→{grade}")
        flash('고장 이력이 수정되었습니다.', 'success')
        if photo_save_failed:
            flash('일부 사진 저장에 실패했습니다. 필요하면 다시 첨부해주세요.', 'warning')
        return redirect(url_for('fault_detail', fault_id=fault_id))

    parts         = db.execute("SELECT * FROM used_parts WHERE fault_id=?", (fault_id,)).fetchall()
    fault_photos  = db.execute(
        "SELECT * FROM fault_photos WHERE fault_id=? AND photo_type='fault' ORDER BY id", (fault_id,)).fetchall()
    action_photos = db.execute(
        "SELECT * FROM fault_photos WHERE fault_id=? AND photo_type='action' ORDER BY id", (fault_id,)).fetchall()
    workers = get_workers(db)
    db.close()
    selected_symptoms = [s.strip() for s in (fault['symptom'] or '').split(',') if s.strip()]
    occurred_at_val = (fault['occurred_at'] or '').replace(' ', 'T')[:16]
    completed_at_val = (fault['completed_at'] or '').replace(' ', 'T')[:16]
    return render_template('fault/edit.html',
                           fault=fault, parts=parts,
                           fault_photos=fault_photos, action_photos=action_photos,
                           symptoms=SYMPTOMS, grades=GRADES,
                           parts_list=PARTS, statuses=STATUSES, workers=workers,
                           selected_symptoms=selected_symptoms,
                           occurred_at_val=occurred_at_val,
                           completed_at_val=completed_at_val,
                           today=date.today().isoformat())


@app.route('/fault/delete/<int:fault_id>', methods=['POST'])
def fault_delete(fault_id):
    if not session.get('edit_authorized'):
        flash('삭제 권한이 없습니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(request.referrer or url_for('fault_list'))
    db = get_db()
    fault = db.execute("SELECT * FROM fault_history WHERE id=?", (fault_id,)).fetchone()
    if not fault:
        db.close()
        abort(404)
    eq_id = fault['equipment_id']
    for fname in [fault['photo_fault'], fault['photo_action']]:
        if fname:
            for d in [FAULT_DIR, ACTION_DIR]:
                fp = os.path.join(d, fname)
                if os.path.exists(fp):
                    os.remove(fp)
    # 다중 사진 파일 삭제
    old_photos = db.execute("SELECT filename FROM fault_photos WHERE fault_id=?", (fault_id,)).fetchall()
    for p in old_photos:
        fp = os.path.join(PHOTOS_DIR, p['filename'])
        if os.path.exists(fp):
            os.remove(fp)
    db.execute("DELETE FROM fault_history WHERE id=?", (fault_id,))
    db.commit()
    db.close()
    log_action('삭제', '고장이력', fault_id, f"{fault['eq_number']} {fault['eq_name']} / {fault['symptom'] or '-'}")
    flash('고장 이력이 삭제되었습니다.', 'warning')
    return redirect(url_for('equipment_detail', eq_id=eq_id))


# ──────────────────────────────────────────────
# 생산 LOT / 설비가동률
# ──────────────────────────────────────────────
@app.route('/production', methods=['GET', 'POST'])
def production_list():
    db = get_db()
    if request.method == 'POST':
        if not session.get('edit_authorized'):
            flash('LOT 등록은 로그인 후 이용해 주세요.', 'danger')
            db.close()
            return redirect(url_for('login', next=url_for('production_list')))
        model_name = request.form.get('model_name', '').strip()
        lot_number = request.form.get('lot_number', '').strip()
        if not model_name or not lot_number:
            flash('모델명과 LOT번호를 입력해 주세요.', 'danger')
            db.close()
            return redirect(url_for('production_list'))
        created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cur = db.execute(
            "INSERT INTO production_lots (model_name,lot_number,started_at,status,started_by) VALUES (?,?,?,?,?)",
            (model_name, lot_number, created_at, '진행중', session.get('actor_name'))
        )
        db.commit()
        log_action('LOT 등록', 'production_lot', cur.lastrowid, f'{model_name} / {lot_number}')
        db.close()
        flash(f'LOT을 등록했습니다: {model_name} / {lot_number}. 각 섹션에서 가동을 시작해 주세요.', 'success')
        return redirect(url_for('production_detail', lot_id=cur.lastrowid))

    lots = [dict(r) for r in db.execute(
        "SELECT * FROM production_lots ORDER BY created_at DESC LIMIT 100"
    ).fetchall()]
    for lot in lots:
        agg = db.execute("""
            SELECT AVG(uptime_rate) as avg_rate, COUNT(*) as n
            FROM production_lot_sections WHERE lot_id=? AND status='완료'
        """, (lot['id'],)).fetchone()
        lot['avg_uptime_rate'] = round(agg['avg_rate'], 1) if agg['avg_rate'] is not None else None
        lot['completed_sections'] = agg['n']

    recent_models = [r['model_name'] for r in db.execute("""
        SELECT model_name, MAX(created_at) as latest
        FROM production_lots GROUP BY model_name ORDER BY latest DESC LIMIT 50
    """).fetchall()]
    db.close()
    return render_template('production/list.html', lots=lots, recent_models=recent_models)


def _compute_section_analysis(db):
    """섹션(SECTIONS) 기준으로 전체 LOT을 가로질러 집계한 가동률 통계.
    완료 여부와 무관하게 모든 가동구간을 포함하며, 진행중인 구간은 현재 시점까지 실시간 계산한다."""
    all_ls = db.execute("""
        SELECT ls.*, l.model_name, l.lot_number
        FROM production_lot_sections ls
        JOIN production_lots l ON ls.lot_id = l.id
        ORDER BY ls.id
    """).fetchall()

    by_section = {s: [] for s in SECTIONS}
    detail_rows = []
    for ls in all_ls:
        stats = compute_section_stats(db, ls['id'], lot_section_row=ls)
        entry = {
            'model_name': ls['model_name'], 'lot_number': ls['lot_number'],
            'status': ls['status'], 'started_at': ls['started_at'], 'ended_at': ls['ended_at'],
            'uptime_rate': stats['uptime_rate'], 'fault_seconds': stats['fault_seconds'],
            'fault_count': stats['fault_count'], 'pause_downtime_seconds': stats['pause_downtime_seconds'],
        }
        by_section.setdefault(ls['section'], []).append(entry)
        detail_rows.append({'section': ls['section'], **entry})

    summary = []
    for section in SECTIONS:
        rows = by_section.get(section, [])
        if rows:
            avg_uptime = round(sum(r['uptime_rate'] for r in rows) / len(rows), 1)
            total_fault_min = sum(r['fault_seconds'] for r in rows) // 60
            total_downtime_min = sum(r['pause_downtime_seconds'] for r in rows) // 60
            total_fault_count = sum(r['fault_count'] for r in rows)
        else:
            avg_uptime = None
            total_fault_min = total_downtime_min = total_fault_count = 0
        summary.append({
            'section': section, 'lot_section_count': len(rows), 'avg_uptime_rate': avg_uptime,
            'total_fault_minutes': total_fault_min, 'total_downtime_minutes': total_downtime_min,
            'total_fault_count': total_fault_count,
        })
    return summary, detail_rows


@app.route('/production/analysis')
def production_analysis():
    db = get_db()
    summary, detail_rows = _compute_section_analysis(db)
    db.close()
    return render_template('production/analysis.html', summary=summary, detail_rows=detail_rows)


@app.route('/production/analysis/export')
def production_analysis_export():
    if not session.get('edit_authorized'):
        flash('엑셀 내보내기는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_analysis')))

    db = get_db()
    summary, detail_rows = _compute_section_analysis(db)
    db.close()

    report_title = '섹션별 가동률 분석 리포트'
    subtitle = f'집계 대상: {len(summary)}개 섹션 · {len(detail_rows)}개 가동구간 (전체 LOT 대상)'
    NCOL = 8
    wb = openpyxl.Workbook()
    ws = xlsx_new_sheet(wb, '섹션별_가동률_분석', first=True)
    row = xlsx_title_block(ws, NCOL, report_title, subtitle)

    rates = [r['avg_uptime_rate'] for r in summary if r['avg_uptime_rate'] is not None]
    overall_rate = round(sum(rates) / len(rates), 1) if rates else None
    total_faults = sum(r['total_fault_count'] for r in summary)
    total_fault_min = sum(r['total_fault_minutes'] for r in summary)
    total_pause_min = sum(r['total_downtime_minutes'] for r in summary)
    total_stop_min = total_fault_min + total_pause_min
    total_sections = sum(r['lot_section_count'] for r in summary)
    # "고장"과 "정지(점검/자재대기 등)"는 표에서는 사유별로 나눠 보여주지만, 결국 둘 다
    # 설비가 멈춘 시간이라는 점에서 같은 범주다. 이를 KPI 카드 맨 앞에 합계로 보여줘서
    # 두 표가 서로 무관한 별개 항목처럼 보이지 않고 "설비정지"라는 하나의 상위 개념 아래
    # 묶여 있다는 걸 한눈에 알 수 있게 한다.
    row = xlsx_kpi_cards(ws, row, NCOL, [
        ('전체 평균가동률', f'{overall_rate}%' if overall_rate is not None else '-',
         _XLSX_COLOR_SUCCESS if (overall_rate is not None and overall_rate >= 95) else
         _XLSX_COLOR_WARNING if (overall_rate is not None and overall_rate >= 85) else
         (_XLSX_COLOR_DANGER if overall_rate is not None else None)),
        ('총 설비정지시간 (고장+정지)', f'{total_stop_min}분', _XLSX_COLOR_DANGER if total_stop_min else None),
        ('└ 고장', f'{total_fault_min}분 ({total_faults}건)', None),
        ('└ 정지 (점검·자재대기 등)', f'{total_pause_min}분', None),
        ('총 가동구간수', f'{total_sections}건', None),
    ])

    row = xlsx_block_heading(ws, row, NCOL, '섹션별 요약')
    summary_rows, fills, point_colors = [], [], []
    for r in summary:
        rate = r['avg_uptime_rate']
        fills.append('E8F5E9' if (rate is not None and rate >= 95) else
                     ('FFFDE0' if (rate is not None and rate >= 85) else 'FFE0E0') if rate is not None else None)
        point_colors.append(
            _XLSX_COLOR_SUCCESS if (rate is not None and rate >= 95) else
            _XLSX_COLOR_WARNING if (rate is not None and rate >= 85) else
            _XLSX_COLOR_DANGER
        )
        summary_rows.append([
            r['section'], r['lot_section_count'], rate if rate is not None else '',
            r['total_fault_minutes'], r['total_fault_count'], r['total_downtime_minutes'],
        ])
    hdr1, last1, row = xlsx_table_at(ws, row,
        ['섹션', '가동구간수', '평균가동률(%)', '총고장시간(분)', '총고장건수', '총정지시간(분)'],
        [14, 12, 14, 14, 12, 14], summary_rows, percent_cols=[3], fill_colors=fills, total_cols=NCOL)
    chart_end1 = xlsx_bar_chart(ws, '섹션별 평균가동률(%)', cat_col=1, data_cols=[3], header_row=hdr1, last_data_row=last1,
                   start_col=1, anchor_row=row, x_title='섹션', y_title='%', y_percent=True,
                   point_colors=point_colors, num_cols=NCOL)
    row = chart_end1 + 2
    chart_end2 = xlsx_bar_chart(ws, '섹션별 총고장/총정지 시간(분)', cat_col=1, data_cols=[4, 6], header_row=hdr1, last_data_row=last1,
                   start_col=1, anchor_row=row, x_title='섹션',
                   series_colors=[_XLSX_COLOR_DANGER, _XLSX_COLOR_WARNING], num_cols=NCOL)
    row = chart_end2 + 2

    row = xlsx_block_heading(ws, row, NCOL, f'섹션별 상세 ({len(detail_rows)}개 가동구간, LOT별)', page_break_before=True)
    detail_table_rows = [[
        r['section'], r['model_name'], r['lot_number'], r['status'],
        r['started_at'], r['ended_at'] or '(진행중)', r['uptime_rate'], r['fault_count'],
    ] for r in detail_rows]
    xlsx_table_at(ws, row, ['섹션', '모델명', 'LOT번호', '상태', '시작', '종료', '가동률(%)', '고장건수'],
                  [12, 16, 14, 10, 18, 18, 12, 10], detail_table_rows, percent_cols=[7], total_cols=NCOL)

    xlsx_finalize_report_sheet(ws, report_title=report_title, tab_color='1F4E79', freeze_row=3)

    log_action('엑셀 내보내기', 'production_analysis', 0, '섹션별 가동률 분석')
    return xlsx_response(wb, '섹션별_가동률_분석')


@app.route('/production/export')
def production_export():
    """생산LOT 설비가동률 종합 리포트. 진행중인 LOT도 현재 시점까지 실시간 계산해서 포함한다."""
    if not session.get('edit_authorized'):
        flash('엑셀 내보내기는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_list')))

    db = get_db()
    lots = db.execute("SELECT * FROM production_lots ORDER BY created_at DESC").fetchall()
    faults = db.execute("""
        SELECT f.*, l.model_name, l.lot_number, ls.section as ls_section
        FROM fault_history f
        JOIN production_lots l ON f.lot_id = l.id
        LEFT JOIN production_lot_sections ls ON f.lot_section_id = ls.id
        ORDER BY f.occurred_at DESC
    """).fetchall()
    pauses = db.execute("""
        SELECT p.*, ls.section, l.model_name, l.lot_number
        FROM production_lot_section_pauses p
        JOIN production_lot_sections ls ON p.lot_section_id = ls.id
        JOIN production_lots l ON ls.lot_id = l.id
        ORDER BY p.paused_at DESC
    """).fetchall()

    report_title = '생산LOT 설비가동률 종합 리포트'
    subtitle = f'집계 대상: {len(lots)}개 LOT · 고장 {len(faults)}건 · 가동정지 {len(pauses)}건'
    NCOL = 10
    wb = openpyxl.Workbook()
    ws = xlsx_new_sheet(wb, '생산LOT_가동률', first=True)
    row = xlsx_title_block(ws, NCOL, report_title, subtitle)

    lot_rows, point_colors = [], []
    for i, lot in enumerate(lots, 1):
        agg = db.execute("""
            SELECT AVG(uptime_rate) as avg_rate, COUNT(*) as n
            FROM production_lot_sections WHERE lot_id=? AND status='완료'
        """, (lot['id'],)).fetchone()
        avg_rate = round(agg['avg_rate'], 1) if agg['avg_rate'] is not None else ''
        point_colors.append(
            _XLSX_COLOR_SUCCESS if (isinstance(avg_rate, (int, float)) and avg_rate >= 95) else
            _XLSX_COLOR_WARNING if (isinstance(avg_rate, (int, float)) and avg_rate >= 85) else
            _XLSX_COLOR_DANGER
        )
        lot_rows.append([i, lot['model_name'], lot['lot_number'], lot['status'],
                         lot['started_at'], lot['ended_at'] or '', avg_rate, agg['n']])

    def _dur_minutes(start_s, end_s):
        if not end_s:
            return None
        return max(0, int((_parse_dt(end_s) - _parse_dt(start_s)).total_seconds() // 60))

    overall_rates = [r[6] for r in lot_rows if isinstance(r[6], (int, float))]
    overall_rate = round(sum(overall_rates) / len(overall_rates), 1) if overall_rates else None
    fault_durations = [d for d in (_dur_minutes(f['occurred_at'], f['completed_at']) for f in faults) if d is not None]
    pause_durations = [d for d in (_dur_minutes(p['paused_at'], p['resumed_at']) for p in pauses) if d is not None]
    mttr = round(sum(fault_durations) / len(fault_durations), 1) if fault_durations else None
    total_fault_min = sum(fault_durations)
    total_pause_min = sum(pause_durations)
    total_stop_min = total_fault_min + total_pause_min
    # 고장·정지(점검/자재대기 등)는 표는 나눠서 보여주되, 결국 둘 다 설비가 멈춘 시간이므로
    # 합계를 맨 앞에 둬서 두 표가 "설비정지"라는 하나의 상위 개념으로 묶여 있음을 보여준다.
    row = xlsx_kpi_cards(ws, row, NCOL, [
        ('전체 평균가동률', f'{overall_rate}%' if overall_rate is not None else '-',
         _XLSX_COLOR_SUCCESS if (overall_rate is not None and overall_rate >= 95) else
         _XLSX_COLOR_WARNING if (overall_rate is not None and overall_rate >= 85) else
         (_XLSX_COLOR_DANGER if overall_rate is not None else None)),
        ('총 설비정지시간 (고장+정지)', f'{total_stop_min}분', _XLSX_COLOR_DANGER if total_stop_min else None),
        ('└ 고장', f'{total_fault_min}분 ({len(faults)}건)', None),
        ('└ 정지 (점검·자재대기 등)', f'{total_pause_min}분', None),
        ('평균 조치시간(MTTR)', f'{mttr}분' if mttr is not None else '-', None),
    ])

    row = xlsx_block_heading(ws, row, NCOL, 'LOT 요약')
    hdr1, last1, row = xlsx_table_at(ws, row,
        ['No', '모델명', 'LOT번호', '상태', '시작', '종료', '평균가동률(%)', '완료섹션수'],
        [5, 16, 14, 10, 18, 18, 14, 12], lot_rows, percent_cols=[7], total_cols=NCOL)
    chart_end = xlsx_bar_chart(ws, 'LOT별 평균가동률(%)', cat_col=3, data_cols=[7], header_row=hdr1, last_data_row=last1,
                   start_col=1, anchor_row=row, x_title='LOT번호', y_title='%', y_percent=True,
                   point_colors=point_colors, num_cols=NCOL)
    row = chart_end + 2

    row = xlsx_block_heading(ws, row, NCOL, '섹션별 상세 (섹션별로 묶어서 표시)', page_break_before=True)
    # LOT을 기준으로 나열하면 같은 섹션의 흐름을 한눈에 보기 어려워서, 섹션(SECTIONS 순서)을
    # 먼저 묶고 그 안에서 LOT별로 나열한다 — "이 섹션이 전체적으로 어떤지"를 바로 볼 수 있게.
    rows_by_section = {s: [] for s in SECTIONS}
    fills_by_section = {s: [] for s in SECTIONS}
    for lot in lots:
        sections = db.execute(
            "SELECT * FROM production_lot_sections WHERE lot_id=? ORDER BY id", (lot['id'],)
        ).fetchall()
        for ls in sections:
            stats = compute_section_stats(db, ls['id'], lot_section_row=ls)
            fill = 'E8F5E9' if stats['uptime_rate'] >= 95 else ('FFFDE0' if stats['uptime_rate'] >= 85 else 'FFE0E0')
            data_row = [
                lot['model_name'], lot['lot_number'], ls['section'], ls['status'],
                ls['started_at'], ls['ended_at'] or '(진행중)',
                stats['available_seconds'] // 60, stats['fault_seconds'] // 60,
                stats['pause_downtime_seconds'] // 60, stats['uptime_rate'],
            ]
            rows_by_section.setdefault(ls['section'], []).append(data_row)
            fills_by_section.setdefault(ls['section'], []).append(fill)

    for section in SECTIONS:
        rows_here = rows_by_section.get(section, [])
        if not rows_here:
            continue
        row = xlsx_subheading(ws, row, NCOL, f'{section} ({len(rows_here)}개 가동구간)')
        _, _, row = xlsx_table_at(ws, row,
            ['모델명', 'LOT번호', '섹션', '상태', '시작', '종료', '가동대상(분)', '고장(분)', '정지-다운타임(분)', '가동률(%)'],
            [16, 14, 12, 10, 18, 18, 12, 10, 16, 12], rows_here, percent_cols=[10],
            fill_colors=fills_by_section.get(section, []), total_cols=NCOL)
    row += 1

    row = xlsx_block_heading(ws, row, NCOL, f'설비정지 내역 ① 고장 ({len(faults)}건) — 발생원인 · 조치시간', page_break_before=True)
    fault_rows = []
    for f in faults:
        occ = _parse_dt(f['occurred_at'])
        comp = _parse_dt(f['completed_at']) if f['completed_at'] else None
        duration = int((comp - occ).total_seconds() // 60) if comp else ''
        fault_rows.append([
            f['model_name'], f['lot_number'], f['ls_section'] or f['section'],
            f['occurred_at'], f['completed_at'] or '(미완료)', duration,
            f['symptom'] or '', f['cause'] or '', f['worker'] or '',
        ])
    hdr3, last3, row = xlsx_table_at(ws, row,
        ['모델명', 'LOT번호', '섹션', '발생시각', '완료시각', '조치시간(분)', '증상', '원인', '조치자'],
        [16, 14, 12, 18, 18, 12, 20, 25, 10], fault_rows, total_cols=NCOL)
    if fault_rows:
        _, row = xlsx_total_row(ws, hdr3 + 1, last3, label_col=1, label=f'합계 ({len(fault_rows)}건)', value_cols=[6])
    row += 1

    row = xlsx_block_heading(ws, row, NCOL, f'설비정지 내역 ② 기타사유 - 점검·자재대기·교체 등 ({len(pauses)}건) — 정지시간 · 정지사유', page_break_before=True)
    pause_rows = []
    for p in pauses:
        paused = _parse_dt(p['paused_at'])
        resumed = _parse_dt(p['resumed_at']) if p['resumed_at'] else None
        duration = int((resumed - paused).total_seconds() // 60) if resumed else ''
        pause_rows.append([
            p['model_name'], p['lot_number'], p['section'],
            p['paused_at'], p['resumed_at'] or '(정지중)', duration, p['reason'] or '퇴근',
        ])
    hdr4, last4, row = xlsx_table_at(ws, row,
        ['모델명', 'LOT번호', '섹션', '정지시각', '재개시각', '정지시간(분)', '사유'],
        [16, 14, 12, 18, 18, 12, 14], pause_rows, total_cols=NCOL)
    if pause_rows:
        xlsx_total_row(ws, hdr4 + 1, last4, label_col=1, label=f'합계 ({len(pause_rows)}건)', value_cols=[6])
    row += 1

    row = _xlsx_add_stop_reason_chart(ws, row, NCOL, total_fault_min, pause_rows, reason_col=7, duration_col=6,
                                    page_break_before=True)

    # 최종 페이지: LOT 하나하나가 아니라 "섹션이 전체적으로 어떤 상태인가"를 전 LOT 통합으로
    # 다시 한번 요약해서, 리포트를 끝까지 본 사람이 결론(섹션별 종합)을 바로 확인할 수 있게 한다.
    analysis_summary, _ = _compute_section_analysis(db)
    row = xlsx_block_heading(ws, row, NCOL, '섹션별 종합 요약 (전체 LOT 통합)', page_break_before=True)
    final_rows, final_fills, final_point_colors = [], [], []
    for r in analysis_summary:
        rate = r['avg_uptime_rate']
        final_fills.append('E8F5E9' if (rate is not None and rate >= 95) else
                           ('FFFDE0' if (rate is not None and rate >= 85) else 'FFE0E0') if rate is not None else None)
        final_point_colors.append(
            _XLSX_COLOR_SUCCESS if (rate is not None and rate >= 95) else
            _XLSX_COLOR_WARNING if (rate is not None and rate >= 85) else
            _XLSX_COLOR_DANGER
        )
        final_rows.append([
            r['section'], r['lot_section_count'], rate if rate is not None else '',
            r['total_fault_minutes'], r['total_fault_count'], r['total_downtime_minutes'],
        ])
    hdr5, last5, row = xlsx_table_at(ws, row,
        ['섹션', '가동구간수', '평균가동률(%)', '총고장시간(분)', '총고장건수', '총정지시간(분)'],
        [14, 12, 14, 14, 12, 14], final_rows, percent_cols=[3], fill_colors=final_fills, total_cols=NCOL)
    row = xlsx_bar_chart(ws, '섹션별 평균가동률(%)', cat_col=1, data_cols=[3], header_row=hdr5, last_data_row=last5,
                   start_col=1, anchor_row=row, x_title='섹션', y_percent=True,
                   point_colors=final_point_colors, num_cols=NCOL) + 1

    xlsx_finalize_report_sheet(ws, report_title=report_title, tab_color='1F4E79', freeze_row=3)

    db.close()
    log_action('엑셀 내보내기', 'production_lot', 0, f'생산LOT 가동률 종합 리포트 ({len(lots)}건)')
    return xlsx_response(wb, '생산LOT_가동률')


@app.route('/production/<int:lot_id>')
def production_detail(lot_id):
    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    if not lot:
        db.close()
        abort(404)

    existing = {r['section']: dict(r) for r in db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? ORDER BY id", (lot_id,)
    ).fetchall()}

    section_rows = []
    for section in SECTIONS:
        ls = existing.get(section)
        if ls:
            if ls['status'] in ('진행중', '일시정지'):
                preview = compute_section_stats(db, ls['id'], lot_section_row=ls)
                ls['uptime_rate'] = preview['uptime_rate']
                ls['available_seconds'] = preview['available_seconds']
                ls['fault_seconds'] = preview['fault_seconds']
                ls['fault_count'] = preview['fault_count']
            start_dt = _parse_dt(ls['started_at'])
            end_dt = _parse_dt(ls['ended_at']) if ls['ended_at'] else datetime.now()
            ls['elapsed_days'] = (end_dt.date() - start_dt.date()).days + 1
            section_rows.append(ls)
        else:
            section_rows.append({'section': section, 'status': '대기'})
    db.close()
    return render_template('production/detail.html', lot=lot, section_rows=section_rows, pause_reasons=PAUSE_REASONS)


@app.route('/production/<int:lot_id>/export')
def production_lot_export(lot_id):
    """LOT 하나의 섹션별 가동률 + 고장/정지 내역 리포트. 진행중인 섹션도 현재 시점까지 실시간 계산."""
    if not session.get('edit_authorized'):
        flash('엑셀 내보내기는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    if not lot:
        db.close()
        abort(404)

    sections = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? ORDER BY id", (lot_id,)
    ).fetchall()

    faults = db.execute("""
        SELECT f.*, ls.section as ls_section
        FROM fault_history f
        LEFT JOIN production_lot_sections ls ON f.lot_section_id = ls.id
        WHERE f.lot_id = ?
        ORDER BY f.occurred_at DESC
    """, (lot_id,)).fetchall()
    pauses = db.execute("""
        SELECT p.*, ls.section
        FROM production_lot_section_pauses p
        JOIN production_lot_sections ls ON p.lot_section_id = ls.id
        WHERE ls.lot_id = ?
        ORDER BY p.paused_at DESC
    """, (lot_id,)).fetchall()

    report_title = f"생산LOT 가동률 리포트 — {lot['model_name']} / {lot['lot_number']}"
    subtitle = (f"모델명: {lot['model_name']}    LOT번호: {lot['lot_number']}    상태: {lot['status']}    "
                f"기간: {lot['started_at']} ~ {lot['ended_at'] or '진행중'}")
    NCOL = 9
    wb = openpyxl.Workbook()
    ws = xlsx_new_sheet(wb, f"{lot['model_name']}_{lot['lot_number']}", first=True)
    row = xlsx_title_block(ws, NCOL, report_title, subtitle)

    section_rows, point_colors, fills = [], [], []
    for ls in sections:
        stats = compute_section_stats(db, ls['id'], lot_section_row=ls)
        rate = stats['uptime_rate']
        fills.append('E8F5E9' if rate >= 95 else ('FFFDE0' if rate >= 85 else 'FFE0E0'))
        point_colors.append(_XLSX_COLOR_SUCCESS if rate >= 95 else (_XLSX_COLOR_WARNING if rate >= 85 else _XLSX_COLOR_DANGER))
        section_rows.append([
            ls['section'], ls['status'], ls['started_at'], ls['ended_at'] or '(진행중)',
            stats['available_seconds'] // 60, stats['fault_seconds'] // 60,
            stats['pause_downtime_seconds'] // 60, rate, stats['fault_count'],
        ])
    db.close()

    def _dur_minutes(start_s, end_s):
        if not end_s:
            return None
        return max(0, int((_parse_dt(end_s) - _parse_dt(start_s)).total_seconds() // 60))

    lot_rate = round(sum(r[7] for r in section_rows) / len(section_rows), 1) if section_rows else None
    fault_durations = [d for d in (_dur_minutes(f['occurred_at'], f['completed_at']) for f in faults) if d is not None]
    pause_durations = [d for d in (_dur_minutes(p['paused_at'], p['resumed_at']) for p in pauses) if d is not None]
    mttr = round(sum(fault_durations) / len(fault_durations), 1) if fault_durations else None
    total_fault_min = sum(fault_durations)
    total_pause_min = sum(pause_durations)
    total_stop_min = total_fault_min + total_pause_min
    row = xlsx_kpi_cards(ws, row, NCOL, [
        ('이 LOT 평균가동률', f'{lot_rate}%' if lot_rate is not None else '-',
         _XLSX_COLOR_SUCCESS if (lot_rate is not None and lot_rate >= 95) else
         _XLSX_COLOR_WARNING if (lot_rate is not None and lot_rate >= 85) else
         (_XLSX_COLOR_DANGER if lot_rate is not None else None)),
        ('총 설비정지시간 (고장+정지)', f'{total_stop_min}분', _XLSX_COLOR_DANGER if total_stop_min else None),
        ('└ 고장', f'{total_fault_min}분 ({len(faults)}건)', None),
        ('└ 정지 (점검·자재대기 등)', f'{total_pause_min}분', None),
        ('평균 조치시간(MTTR)', f'{mttr}분' if mttr is not None else '-', None),
    ])

    row = xlsx_block_heading(ws, row, NCOL, '섹션별 가동률 요약')
    hdr1, last1, row = xlsx_table_at(ws, row,
        ['섹션', '상태', '시작', '종료', '가동대상(분)', '고장(분)', '정지-다운타임(분)', '가동률(%)', '고장건수'],
        [12, 10, 18, 18, 12, 10, 16, 12, 10], section_rows, percent_cols=[8], fill_colors=fills, total_cols=NCOL)
    chart_end1 = xlsx_bar_chart(ws, '섹션별 가동률(%)', cat_col=1, data_cols=[8], header_row=hdr1, last_data_row=last1,
                   start_col=1, anchor_row=row, x_title='섹션', y_title='%', y_percent=True,
                   point_colors=point_colors, num_cols=NCOL)
    row = chart_end1 + 2
    chart_end2 = xlsx_bar_chart(ws, '섹션별 고장/정지 시간(분)', cat_col=1, data_cols=[6, 7], header_row=hdr1, last_data_row=last1,
                   start_col=1, anchor_row=row, x_title='섹션',
                   series_colors=[_XLSX_COLOR_DANGER, _XLSX_COLOR_WARNING], num_cols=NCOL)
    row = chart_end2 + 2

    row = xlsx_block_heading(ws, row, NCOL, f'설비정지 내역 ① 고장 ({len(faults)}건) — 발생원인 · 조치시간', page_break_before=True)
    fault_rows = []
    for f in faults:
        occ = _parse_dt(f['occurred_at'])
        comp = _parse_dt(f['completed_at']) if f['completed_at'] else None
        duration = int((comp - occ).total_seconds() // 60) if comp else ''
        fault_rows.append([
            f['ls_section'] or f['section'], f['occurred_at'], f['completed_at'] or '(미완료)',
            duration, f['symptom'] or '', f['cause'] or '', f['worker'] or '',
        ])
    hdr2, last2, row = xlsx_table_at(ws, row,
        ['섹션', '발생시각', '완료시각', '조치시간(분)', '증상', '원인', '조치자'],
        [12, 18, 18, 12, 20, 25, 10], fault_rows, total_cols=NCOL)
    if fault_rows:
        _, row = xlsx_total_row(ws, hdr2 + 1, last2, label_col=1, label=f'합계 ({len(fault_rows)}건)', value_cols=[4])
    row += 1

    row = xlsx_block_heading(ws, row, NCOL, f'설비정지 내역 ② 기타사유 - 점검·자재대기·교체 등 ({len(pauses)}건) — 정지시간 · 정지사유', page_break_before=True)
    pause_rows = []
    for p in pauses:
        paused = _parse_dt(p['paused_at'])
        resumed = _parse_dt(p['resumed_at']) if p['resumed_at'] else None
        duration = int((resumed - paused).total_seconds() // 60) if resumed else ''
        pause_rows.append([p['section'], p['paused_at'], p['resumed_at'] or '(정지중)', duration, p['reason'] or '퇴근'])
    hdr3, last3, row = xlsx_table_at(ws, row, ['섹션', '정지시각', '재개시각', '정지시간(분)', '사유'],
                                      [12, 18, 18, 12, 14], pause_rows, total_cols=NCOL)
    if pause_rows:
        xlsx_total_row(ws, hdr3 + 1, last3, label_col=1, label=f'합계 ({len(pause_rows)}건)', value_cols=[4])
    row += 1

    row = _xlsx_add_stop_reason_chart(ws, row, NCOL, total_fault_min, pause_rows, reason_col=5, duration_col=4,
                                    page_break_before=True)

    xlsx_finalize_report_sheet(ws, report_title=report_title, tab_color='1F4E79', freeze_row=3)

    log_action('엑셀 내보내기', 'production_lot', lot_id, f"{lot['model_name']} / {lot['lot_number']}")
    return xlsx_response(wb, f"{lot['model_name']}_{lot['lot_number']}_가동률")


@app.route('/production/<int:lot_id>/sections/start-all', methods=['POST'])
def production_sections_start_all(lot_id):
    if not session.get('edit_authorized'):
        flash('가동 시작은 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    if not lot:
        db.close()
        abort(404)

    busy_sections = {r['section'] for r in db.execute(
        "SELECT section FROM production_lot_sections WHERE status IN ('진행중','일시정지')"
    ).fetchall()}
    already_used = {r['section'] for r in db.execute(
        "SELECT section FROM production_lot_sections WHERE lot_id=?", (lot_id,)
    ).fetchall()}
    to_start = [s for s in SECTIONS if s not in busy_sections and s not in already_used]

    started_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    started = []
    skipped = []
    for section in to_start:
        try:
            db.execute(
                "INSERT INTO production_lot_sections (lot_id,section,started_at,status) VALUES (?,?,?,?)",
                (lot_id, section, started_at, '진행중')
            )
            started.append(section)
        except sqlite3.IntegrityError:
            # 동시 요청으로 이미 다른 곳에서 시작된 섹션 (DB 유니크 제약이 최종 방어)
            skipped.append(section)
    db.commit()
    if started:
        log_action('섹션 일괄시작', 'production_lot', lot_id, ', '.join(started))
        msg = f"{len(started)}개 섹션의 가동을 시작했습니다 ({', '.join(started)})."
        if skipped:
            msg += f" (동시에 시작되어 건너뜀: {', '.join(skipped)})"
        flash(msg, 'success')
    else:
        flash('시작할 수 있는 섹션이 없습니다 (이미 모두 진행중이거나 이 LOT에서 사용된 섹션입니다).', 'warning')
    db.close()
    return redirect(url_for('production_detail', lot_id=lot_id))


@app.route('/production/<int:lot_id>/sections/pause-all', methods=['POST'])
def production_sections_pause_all(lot_id):
    if not session.get('edit_authorized'):
        flash('가동 정지는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    reason_sel = request.form.get('reason', '').strip()
    reason_custom = request.form.get('reason_custom', '').strip()
    reason = (reason_custom if reason_sel == '기타' else reason_sel) or '퇴근'

    running = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND status='진행중'", (lot_id,)
    ).fetchall()
    paused_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for ls in running:
        db.execute(
            "INSERT INTO production_lot_section_pauses (lot_section_id, paused_at, reason) VALUES (?,?,?)",
            (ls['id'], paused_at, reason)
        )
        db.execute("UPDATE production_lot_sections SET status='일시정지' WHERE id=?", (ls['id'],))
    db.commit()
    sections = [r['section'] for r in running]
    if sections:
        log_action('섹션 일괄정지', 'production_lot', lot_id, f"{', '.join(sections)} ({reason})")
        flash(f"{len(sections)}개 섹션의 가동을 정지했습니다 ({', '.join(sections)}).", 'success')
    else:
        flash('정지할 진행중인 섹션이 없습니다.', 'warning')
    db.close()
    return redirect(url_for('production_detail', lot_id=lot_id))


@app.route('/production/<int:lot_id>/section/<section>/start', methods=['POST'])
def production_section_start(lot_id, section):
    if not session.get('edit_authorized'):
        flash('가동 시작은 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))
    if section not in SECTIONS:
        abort(404)

    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    if not lot:
        db.close()
        abort(404)
    if db.execute(
        "SELECT id FROM production_lot_sections WHERE section=? AND status IN ('진행중','일시정지')", (section,)
    ).fetchone():
        db.close()
        flash(f'{section}은(는) 이미 다른 가동(또는 정지 중인 가동)이 있습니다.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    started_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        db.execute(
            "INSERT INTO production_lot_sections (lot_id,section,started_at,status) VALUES (?,?,?,?)",
            (lot_id, section, started_at, '진행중')
        )
        db.commit()
    except sqlite3.IntegrityError:
        # 동시 요청으로 이미 다른 곳에서 이 섹션을 시작시킨 경우 (DB 유니크 제약이 최종 방어)
        db.close()
        flash(f'{section}은(는) 방금 다른 곳에서 이미 가동이 시작되었습니다.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))
    log_action('섹션 가동시작', 'production_lot', lot_id, f"{section} / {lot['model_name']} {lot['lot_number']}")
    db.close()
    flash(f'{section} 가동을 시작했습니다.', 'success')
    return redirect(url_for('production_detail', lot_id=lot_id))


@app.route('/production/<int:lot_id>/section/<section>/pause', methods=['POST'])
def production_section_pause(lot_id, section):
    if not session.get('edit_authorized'):
        flash('가동 정지는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    ls = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND section=? AND status='진행중'",
        (lot_id, section)
    ).fetchone()
    if not ls:
        db.close()
        flash('진행중인 가동 구간을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    reason_sel = request.form.get('reason', '').strip()
    reason_custom = request.form.get('reason_custom', '').strip()
    reason = reason_custom if reason_sel == '기타' else reason_sel
    reason = reason or '퇴근'

    paused_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "INSERT INTO production_lot_section_pauses (lot_section_id, paused_at, reason) VALUES (?,?,?)",
        (ls['id'], paused_at, reason)
    )
    db.execute("UPDATE production_lot_sections SET status='일시정지' WHERE id=?", (ls['id'],))
    db.commit()
    log_action('섹션 가동정지', 'production_lot', lot_id, f"{section} ({reason})")
    db.close()
    flash(f'{section} 가동을 정지했습니다. (다음 출근 시 재개하세요)', 'success')
    return redirect(url_for('production_detail', lot_id=lot_id))


@app.route('/production/<int:lot_id>/section/<section>/resume', methods=['POST'])
def production_section_resume(lot_id, section):
    if not session.get('edit_authorized'):
        flash('가동 재개는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    ls = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND section=? AND status='일시정지'",
        (lot_id, section)
    ).fetchone()
    if not ls:
        db.close()
        flash('정지중인 가동 구간을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    resumed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "UPDATE production_lot_section_pauses SET resumed_at=? WHERE lot_section_id=? AND resumed_at IS NULL",
        (resumed_at, ls['id'])
    )
    db.execute("UPDATE production_lot_sections SET status='진행중' WHERE id=?", (ls['id'],))
    db.commit()
    log_action('섹션 가동재개', 'production_lot', lot_id, f"{section}")
    db.close()
    flash(f'{section} 가동을 재개했습니다.', 'success')
    return redirect(url_for('production_detail', lot_id=lot_id))


@app.route('/production/<int:lot_id>/section/<section>/end', methods=['POST'])
def production_section_end(lot_id, section):
    if not session.get('edit_authorized'):
        flash('가동 종료는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    ls = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND section=? AND status IN ('진행중','일시정지')",
        (lot_id, section)
    ).fetchone()
    if not ls:
        db.close()
        flash('진행중인 가동 구간을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    ended_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "UPDATE production_lot_section_pauses SET resumed_at=? WHERE lot_section_id=? AND resumed_at IS NULL",
        (ended_at, ls['id'])
    )
    ls_dict = dict(ls)
    ls_dict['ended_at'] = ended_at
    stats = compute_section_stats(db, ls['id'], lot_section_row=ls_dict)
    db.execute("""
        UPDATE production_lot_sections
        SET ended_at=?, status='완료', available_seconds=?, fault_seconds=?, fault_count=?,
            pause_downtime_seconds=?, uptime_rate=?
        WHERE id=?
    """, (ended_at, stats['available_seconds'], stats['fault_seconds'], stats['fault_count'],
          stats['pause_downtime_seconds'], stats['uptime_rate'], ls['id']))
    db.commit()
    log_action('섹션 가동종료', 'production_lot', lot_id, f"{section} / 가동률 {stats['uptime_rate']}%")
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    db.close()
    if lot:
        notify_section_end(lot, section, stats)
    flash(f"{section} 가동을 종료했습니다. (가동률 {stats['uptime_rate']}%)", 'success')
    return redirect(url_for('production_detail', lot_id=lot_id))


@app.route('/production/<int:lot_id>/section/<section>/edit_times', methods=['POST'])
def production_section_edit_times(lot_id, section):
    """가동 시작/종료 시각을 실제 시각에 맞게 수정. 이미 종료된 구간이면 수정된 시각 기준으로
    가동대상시간/고장시간/가동률을 다시 계산해서 저장한다 (진행중인 구간은 매번 실시간 계산되므로
    시작시각만 바꿔도 다음 조회 때 자동 반영됨)."""
    if not session.get('edit_authorized'):
        flash('시각 수정은 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    ls = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND section=? ORDER BY id DESC LIMIT 1",
        (lot_id, section)
    ).fetchone()
    if not ls:
        db.close()
        flash('가동 구간을 찾을 수 없습니다.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    started_at = parse_optional_datetime(request.form.get('started_at', ''))
    if not started_at:
        db.close()
        flash('시작 시각을 올바르게 입력해 주세요.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    ended_at = ls['ended_at']
    if ls['status'] == '완료':
        ended_at = parse_optional_datetime(request.form.get('ended_at', ''))
        if not ended_at:
            db.close()
            flash('종료 시각을 올바르게 입력해 주세요.', 'danger')
            return redirect(url_for('production_detail', lot_id=lot_id))
        if ended_at <= started_at:
            db.close()
            flash('종료 시각은 시작 시각보다 늦어야 합니다.', 'danger')
            return redirect(url_for('production_detail', lot_id=lot_id))

    db.execute("UPDATE production_lot_sections SET started_at=?, ended_at=? WHERE id=?",
               (started_at, ended_at, ls['id']))

    if ls['status'] == '완료':
        ls_dict = dict(ls)
        ls_dict['started_at'] = started_at
        ls_dict['ended_at'] = ended_at
        stats = compute_section_stats(db, ls['id'], lot_section_row=ls_dict)
        db.execute("""
            UPDATE production_lot_sections
            SET available_seconds=?, fault_seconds=?, fault_count=?, pause_downtime_seconds=?, uptime_rate=?
            WHERE id=?
        """, (stats['available_seconds'], stats['fault_seconds'], stats['fault_count'],
              stats['pause_downtime_seconds'], stats['uptime_rate'], ls['id']))
        detail = f"{section} 시작:{started_at} 종료:{ended_at} / 가동률 재계산 {stats['uptime_rate']}%"
    else:
        detail = f"{section} 시작:{started_at}"

    db.commit()
    log_action('섹션 시각 수정', 'production_lot', lot_id, detail)
    db.close()
    flash(f'{section}의 가동 시각을 수정했습니다.', 'success')
    return redirect(url_for('production_detail', lot_id=lot_id))


def _serialize_timeline(db, ls):
    win_start = _parse_dt(ls['started_at'])
    win_end = _parse_dt(ls['ended_at']) if ls['ended_at'] else datetime.now()
    total = max(1.0, (win_end - win_start).total_seconds())
    segments = compute_section_timeline(db, ls)
    seg_out = []
    for s in segments:
        dur = (s['end'] - s['start']).total_seconds()
        seg_out.append({
            'type': s['type'],
            'start': s['start'].strftime('%Y-%m-%d %H:%M:%S'),
            'end': s['end'].strftime('%Y-%m-%d %H:%M:%S'),
            'label': s['label'],
            'duration_seconds': int(dur),
            'pct': round(dur / total * 100, 3),
        })
    return {
        'window_start': win_start.strftime('%Y-%m-%d %H:%M:%S'),
        'window_end': win_end.strftime('%Y-%m-%d %H:%M:%S'),
        'segments': seg_out,
    }


@app.route('/production/<int:lot_id>/section/<section>')
def production_section_detail(lot_id, section):
    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    ls = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND section=? ORDER BY id DESC LIMIT 1",
        (lot_id, section)
    ).fetchone()
    if not lot or not ls:
        db.close()
        abort(404)
    stats = compute_section_stats(db, ls['id'], lot_section_row=ls)
    timeline = _serialize_timeline(db, ls)
    db.close()
    return render_template('production/section_detail.html', lot=lot, ls=ls, stats=stats,
                           timeline=timeline)


@app.route('/production/<int:lot_id>/section/<section>/export')
def production_section_export(lot_id, section):
    """섹션 가동구간 하나의 상세 리포트 (가동률 요약 + 고장 내역 + 가동정지 내역)."""
    if not session.get('edit_authorized'):
        flash('엑셀 내보내기는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_section_detail', lot_id=lot_id, section=section)))

    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    ls = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND section=? ORDER BY id DESC LIMIT 1",
        (lot_id, section)
    ).fetchone()
    if not lot or not ls:
        db.close()
        abort(404)
    stats = compute_section_stats(db, ls['id'], lot_section_row=ls)
    db.close()

    report_title = f"섹션 가동률 리포트 — {lot['model_name']} / {lot['lot_number']} / {section}"
    rate = stats['uptime_rate']
    subtitle = (f"모델명: {lot['model_name']}    LOT번호: {lot['lot_number']}    섹션: {section}    "
                f"상태: {ls['status']}    기간: {ls['started_at']} ~ {ls['ended_at'] or '진행중'}    "
                f"가동률: {rate}%")
    NCOL = 6
    wb = openpyxl.Workbook()
    ws = xlsx_new_sheet(wb, f"{section}_가동률", first=True)
    row = xlsx_title_block(ws, NCOL, report_title, subtitle)

    completed_fault_durations = [f['duration_minutes'] for f in stats['faults'] if not f['ongoing']]
    mttr = round(sum(completed_fault_durations) / len(completed_fault_durations), 1) if completed_fault_durations else None
    total_stop_min = (stats['fault_seconds'] + stats['pause_downtime_seconds']) // 60
    row = xlsx_kpi_cards(ws, row, NCOL, [
        ('가동률', f'{rate}%',
         _XLSX_COLOR_SUCCESS if rate >= 95 else (_XLSX_COLOR_WARNING if rate >= 85 else _XLSX_COLOR_DANGER)),
        ('총 설비정지시간 (고장+정지)', f'{total_stop_min}분', _XLSX_COLOR_DANGER if total_stop_min else None),
        ('└ 고장', f"{stats['fault_seconds'] // 60}분 ({stats['fault_count']}건)", None),
        ('└ 정지 (점검·자재대기 등)', f"{stats['pause_downtime_seconds'] // 60}분", None),
        ('평균 조치시간(MTTR)', f'{mttr}분' if mttr is not None else '-', None),
    ])

    row = xlsx_block_heading(ws, row, NCOL, '가동률 요약 지표')
    fill = 'E8F5E9' if rate >= 95 else ('FFFDE0' if rate >= 85 else 'FFE0E0')
    hdr0, last0, row = xlsx_table_at(ws, row,
        ['가동대상(분)', '고장(분)', '정지-다운타임(분)', '가동률(%)', '고장건수'],
        [16, 12, 18, 12, 12],
        [[stats['available_seconds'] // 60, stats['fault_seconds'] // 60,
          stats['pause_downtime_seconds'] // 60, rate, stats['fault_count']]],
        percent_cols=[4], fill_colors=[fill], total_cols=NCOL)
    row += 1

    row = xlsx_block_heading(ws, row, NCOL, f"설비정지 내역 ① 고장 ({len(stats['faults'])}건) — 발생원인 · 조치시간", page_break_before=True)
    fault_rows = [[
        f['occurred_at'], '(미완료)' if f['ongoing'] else f['completed_at'],
        f['duration_minutes'], f['symptom'] or '', f['cause'] or '', f['worker'] or '',
    ] for f in stats['faults']]
    hdr1, last1, row = xlsx_table_at(ws, row, ['발생시각', '완료시각', '조치시간(분)', '증상', '원인', '조치자'],
                                      [18, 18, 12, 20, 25, 10], fault_rows, total_cols=NCOL)
    if fault_rows:
        _, row = xlsx_total_row(ws, hdr1 + 1, last1, label_col=1, label=f'합계 ({len(fault_rows)}건)', value_cols=[3])
    row += 1

    row = xlsx_block_heading(ws, row, NCOL, f"설비정지 내역 ② 기타사유 - 점검·자재대기·교체 등 ({len(stats['pauses'])}건) — 정지시간 · 정지사유", page_break_before=True)
    pause_rows = [[
        p['paused_at'], '(정지중)' if p['ongoing'] else p['resumed_at'], p['duration_minutes'], p['reason'],
    ] for p in stats['pauses']]
    hdr2, last2, row = xlsx_table_at(ws, row, ['정지시각', '재개시각', '정지시간(분)', '사유'],
                                      [18, 18, 12, 14], pause_rows, total_cols=NCOL)
    if pause_rows:
        _, row = xlsx_total_row(ws, hdr2 + 1, last2, label_col=1, label=f'합계 ({len(pause_rows)}건)', value_cols=[3])
    row += 1

    # 일별 추이: 고장/가동정지를 날짜별로 묶어서 시간대별 그래프로 표시
    daily = {}
    for f in stats['faults']:
        d = f['occurred_at'][:10]
        daily.setdefault(d, {'fault': 0, 'downtime': 0})
        daily[d]['fault'] += f['duration_minutes']
    for p in stats['pauses']:
        d = p['paused_at'][:10]
        daily.setdefault(d, {'fault': 0, 'downtime': 0})
        daily[d]['downtime'] += p['duration_minutes']

    row = xlsx_block_heading(ws, row, NCOL, '일별 추이', page_break_before=True)
    daily_rows = [[d, daily[d]['fault'], daily[d]['downtime']] for d in sorted(daily.keys())]
    hdr3, last3, row = xlsx_table_at(ws, row, ['날짜', '고장(분)', '정지(분)'], [14, 12, 12], daily_rows, total_cols=NCOL)
    row = xlsx_bar_chart(ws, '일별 고장/정지 시간(분)', cat_col=1, data_cols=[2, 3], header_row=hdr3, last_data_row=last3,
                   start_col=1, anchor_row=row, x_title='날짜',
                   series_colors=[_XLSX_COLOR_DANGER, _XLSX_COLOR_WARNING], num_cols=NCOL) + 1

    row = _xlsx_add_stop_reason_chart(ws, row, NCOL, stats['fault_seconds'] // 60, pause_rows, reason_col=4, duration_col=3,
                                    page_break_before=True)

    xlsx_finalize_report_sheet(ws, report_title=report_title, tab_color='1F4E79', freeze_row=3)

    log_action('엑셀 내보내기', 'production_lot_section', ls['id'],
               f"{lot['model_name']} / {lot['lot_number']} / {section}")
    return xlsx_response(wb, f"{lot['model_name']}_{lot['lot_number']}_{section}_가동률")


@app.route('/production/<int:lot_id>/section/<section>/timeline.json')
def production_section_timeline_json(lot_id, section):
    db = get_db()
    ls = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND section=? ORDER BY id DESC LIMIT 1",
        (lot_id, section)
    ).fetchone()
    if not ls:
        db.close()
        return jsonify(ok=False, error='not found'), 404
    stats = compute_section_stats(db, ls['id'], lot_section_row=ls)
    timeline = _serialize_timeline(db, ls)
    db.close()
    return jsonify(ok=True, status=ls['status'], uptime_rate=stats['uptime_rate'],
                   fault_count=stats['fault_count'], fault_seconds=stats['fault_seconds'],
                   pause_count=stats['pause_count'], pause_downtime_seconds=stats['pause_downtime_seconds'],
                   available_seconds=stats['available_seconds'], **timeline)


@app.route('/production/<int:lot_id>/end', methods=['POST'])
def production_end(lot_id):
    if not session.get('edit_authorized'):
        flash('LOT 종료는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_detail', lot_id=lot_id)))

    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    if not lot or lot['status'] == '완료':
        db.close()
        flash('이미 종료되었거나 존재하지 않는 LOT입니다.', 'danger')
        return redirect(url_for('production_list'))

    # 아직 진행중/정지중인 섹션은 지금 시각으로 강제 종료 처리
    running = db.execute(
        "SELECT * FROM production_lot_sections WHERE lot_id=? AND status IN ('진행중','일시정지')", (lot_id,)
    ).fetchall()
    ended_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    for ls in running:
        db.execute(
            "UPDATE production_lot_section_pauses SET resumed_at=? WHERE lot_section_id=? AND resumed_at IS NULL",
            (ended_at, ls['id'])
        )
        ls_dict = dict(ls)
        ls_dict['ended_at'] = ended_at
        stats = compute_section_stats(db, ls['id'], lot_section_row=ls_dict)
        db.execute("""
            UPDATE production_lot_sections
            SET ended_at=?, status='완료', available_seconds=?, fault_seconds=?, fault_count=?,
                pause_downtime_seconds=?, uptime_rate=?
            WHERE id=?
        """, (ended_at, stats['available_seconds'], stats['fault_seconds'], stats['fault_count'],
              stats['pause_downtime_seconds'], stats['uptime_rate'], ls['id']))

    db.execute("UPDATE production_lots SET ended_at=?, status='완료' WHERE id=?", (ended_at, lot_id))
    db.commit()
    log_action('LOT 종료', 'production_lot', lot_id, f"{lot['model_name']} / {lot['lot_number']}")
    db.close()
    flash('LOT을 종료했습니다.', 'success')
    return redirect(url_for('production_detail', lot_id=lot_id))


@app.route('/production/<int:lot_id>/edit', methods=['GET', 'POST'])
def production_edit(lot_id):
    if not session.get('edit_authorized'):
        flash('수정 권한이 없습니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    db = get_db()
    lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
    if not lot:
        db.close()
        abort(404)

    if request.method == 'POST':
        model_name = request.form.get('model_name', '').strip()
        lot_number = request.form.get('lot_number', '').strip()
        if not model_name or not lot_number:
            db.close()
            flash('모델명과 LOT번호를 입력해 주세요.', 'danger')
            return redirect(url_for('production_edit', lot_id=lot_id))
        db.execute("UPDATE production_lots SET model_name=?, lot_number=? WHERE id=?",
                   (model_name, lot_number, lot_id))
        db.commit()
        log_action('LOT 수정', 'production_lot', lot_id, f'{model_name} / {lot_number}')
        db.close()
        flash('LOT 정보를 수정했습니다.', 'success')
        return redirect(url_for('production_detail', lot_id=lot_id))

    db.close()
    return render_template('production/edit.html', lot=lot)


@app.route('/production/<int:lot_id>/delete', methods=['POST'])
def production_delete(lot_id):
    if not session.get('edit_authorized'):
        flash('삭제 권한이 없습니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(url_for('production_detail', lot_id=lot_id))

    db = get_db()
    try:
        lot = db.execute("SELECT * FROM production_lots WHERE id=?", (lot_id,)).fetchone()
        if not lot:
            abort(404)
        db.execute("UPDATE fault_history SET lot_id=NULL, lot_section_id=NULL WHERE lot_id=?", (lot_id,))
        db.execute("""
            DELETE FROM production_lot_section_pauses
            WHERE lot_section_id IN (SELECT id FROM production_lot_sections WHERE lot_id=?)
        """, (lot_id,))
        db.execute("DELETE FROM production_lot_section_stats WHERE lot_id=?", (lot_id,))
        db.execute("DELETE FROM production_lot_sections WHERE lot_id=?", (lot_id,))
        db.execute("DELETE FROM production_lots WHERE id=?", (lot_id,))
        db.commit()
    finally:
        db.close()
    log_action('LOT 삭제', 'production_lot', lot_id, f"{lot['model_name']} / {lot['lot_number']}")
    flash('LOT을 삭제했습니다.', 'warning')
    return redirect(url_for('production_list'))


@app.route('/production/settings', methods=['GET', 'POST'])
def production_settings():
    if not session.get('edit_authorized'):
        flash('설정 변경은 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('production_settings')))

    db = get_db()
    if request.method == 'POST':
        schedules = db.execute("SELECT id, name FROM break_schedules").fetchall()
        for sch in schedules:
            start = request.form.get(f'start_{sch["id"]}', '').strip()
            end = request.form.get(f'end_{sch["id"]}', '').strip()
            if start and end:
                db.execute("UPDATE break_schedules SET start_time=?, end_time=? WHERE id=?",
                           (start, end, sch['id']))
        db.commit()
        db.close()
        flash('휴식시간 설정을 저장했습니다.', 'success')
        return redirect(url_for('production_settings'))

    schedules = db.execute("SELECT * FROM break_schedules ORDER BY id").fetchall()
    db.close()
    return render_template('production/settings.html', schedules=schedules)


# ──────────────────────────────────────────────
# 분석
# ──────────────────────────────────────────────
def _compute_equipment_reliability(db):
    """설비별 MTBF(평균고장간격, 일)/MTTR(평균수리시간, 분).
    MTBF는 같은 설비에서 연속된 고장 발생시각 사이의 평균 간격 (고장이 2건 이상이어야 계산 가능).
    MTTR은 조치완료된 고장들의 평균 소요시간 (완료되지 않은 고장은 제외)."""
    rows = db.execute("""
        SELECT equipment_id, occurred_at, completed_at
        FROM fault_history
        WHERE equipment_id IS NOT NULL
        ORDER BY equipment_id, occurred_at
    """).fetchall()
    by_eq = {}
    for r in rows:
        by_eq.setdefault(r['equipment_id'], []).append(r)

    result = {}
    for eq_id, faults in by_eq.items():
        occurred_times = [_parse_dt(f['occurred_at']) for f in faults]
        if len(occurred_times) >= 2:
            gaps = [(occurred_times[i] - occurred_times[i - 1]).total_seconds() / 86400
                    for i in range(1, len(occurred_times))]
            mtbf_days = sum(gaps) / len(gaps)
        else:
            mtbf_days = None

        durations = []
        for f in faults:
            if f['completed_at']:
                occ = _parse_dt(f['occurred_at'])
                comp = _parse_dt(f['completed_at'])
                if comp > occ:
                    durations.append((comp - occ).total_seconds() / 60)
        mttr_minutes = sum(durations) / len(durations) if durations else None

        result[eq_id] = {
            'mtbf_days': round(mtbf_days, 1) if mtbf_days is not None else None,
            'mttr_minutes': round(mttr_minutes, 1) if mttr_minutes is not None else None,
        }
    return result


@app.route('/analysis/repeat')
def analysis_repeat():
    db = get_db()
    eq_stat_rows = db.execute("""
        SELECT f.equipment_id, f.eq_number, f.eq_name, e.section,
               COUNT(*) as total,
               SUM(CASE WHEN f.grade='A' THEN 1 ELSE 0 END) as grade_a,
               SUM(CASE WHEN f.grade='B' THEN 1 ELSE 0 END) as grade_b,
               SUM(CASE WHEN f.grade='C' THEN 1 ELSE 0 END) as grade_c
        FROM fault_history f
        LEFT JOIN equipment e ON f.equipment_id = e.id
        GROUP BY f.equipment_id ORDER BY total DESC
    """).fetchall()
    reliability = _compute_equipment_reliability(db)
    eq_stat = []
    for row in eq_stat_rows:
        d = dict(row)
        d.update(reliability.get(row['equipment_id'], {'mtbf_days': None, 'mttr_minutes': None}))
        eq_stat.append(d)

    sym_stat = db.execute("""
        SELECT symptom, COUNT(*) as total FROM fault_history
        GROUP BY symptom ORDER BY total DESC
    """).fetchall()

    parts_stat = db.execute("""
        SELECT part_name, SUM(quantity) as total, COUNT(DISTINCT fault_id) as cases
        FROM used_parts GROUP BY part_name ORDER BY total DESC
    """).fetchall()

    grade_stat = db.execute("""
        SELECT grade, COUNT(*) as total FROM fault_history
        GROUP BY grade ORDER BY grade
    """).fetchall()

    db.close()
    return render_template('analysis/repeat.html',
                           eq_stat=eq_stat, sym_stat=sym_stat,
                           parts_stat=parts_stat, grade_stat=grade_stat)


@app.route('/analysis/repeat/export')
def analysis_repeat_export():
    if not session.get('edit_authorized'):
        flash('엑셀 내보내기는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('analysis_repeat')))

    db = get_db()
    eq_stat_rows = db.execute("""
        SELECT f.equipment_id, f.eq_number, f.eq_name, e.section,
               COUNT(*) as total,
               SUM(CASE WHEN f.grade='A' THEN 1 ELSE 0 END) as grade_a,
               SUM(CASE WHEN f.grade='B' THEN 1 ELSE 0 END) as grade_b,
               SUM(CASE WHEN f.grade='C' THEN 1 ELSE 0 END) as grade_c
        FROM fault_history f
        LEFT JOIN equipment e ON f.equipment_id = e.id
        GROUP BY f.equipment_id ORDER BY total DESC
    """).fetchall()
    reliability = _compute_equipment_reliability(db)
    sym_stat = db.execute("""
        SELECT symptom, COUNT(*) as total FROM fault_history
        GROUP BY symptom ORDER BY total DESC
    """).fetchall()
    parts_stat = db.execute("""
        SELECT part_name, SUM(quantity) as total, COUNT(DISTINCT fault_id) as cases
        FROM used_parts GROUP BY part_name ORDER BY total DESC
    """).fetchall()
    grade_stat = db.execute("""
        SELECT grade, COUNT(*) as total FROM fault_history
        GROUP BY grade ORDER BY grade
    """).fetchall()
    db.close()

    wb = openpyxl.Workbook()

    ws1 = xlsx_sheet(wb, '설비별 고장현황',
                      ['순위', '설비번호', '설비명', '섹션', '합계', 'A 생산정지', 'B 품질영향', 'C 경미',
                       'MTBF(일)', 'MTTR(분)'],
                      [6, 14, 20, 12, 8, 12, 12, 10, 10, 10], first=True)
    for i, row in enumerate(eq_stat_rows, 1):
        rel = reliability.get(row['equipment_id'], {'mtbf_days': None, 'mttr_minutes': None})
        xlsx_row(ws1, i + 1, [i, row['eq_number'], row['eq_name'], row['section'] or '',
                              row['total'], row['grade_a'] or 0, row['grade_b'] or 0, row['grade_c'] or 0,
                              rel['mtbf_days'] if rel['mtbf_days'] is not None else '',
                              rel['mttr_minutes'] if rel['mttr_minutes'] is not None else ''])

    ws2 = xlsx_sheet(wb, '증상별 발생현황', ['순위', '증상', '건수'], [6, 30, 10])
    for i, row in enumerate(sym_stat, 1):
        xlsx_row(ws2, i + 1, [i, row['symptom'] or '미입력', row['total']])

    ws3 = xlsx_sheet(wb, '부품별 교체현황', ['순위', '부품명', '교체건수', '총수량'], [6, 25, 12, 12])
    for i, row in enumerate(parts_stat, 1):
        xlsx_row(ws3, i + 1, [i, row['part_name'], row['cases'], row['total']])

    ws4 = xlsx_sheet(wb, '고장등급 분포', ['등급', '건수'], [10, 10])
    for i, row in enumerate(grade_stat, 1):
        xlsx_row(ws4, i + 1, [row['grade'] or 'D', row['total']])

    log_action('엑셀 내보내기', 'analysis', 0, '반복고장 분석')
    return xlsx_response(wb, '반복고장분석')


@app.route('/analysis/trend')
def analysis_trend():
    db = get_db()
    period = request.args.get('period', 'monthly')
    section = request.args.get('section', '')
    trend, trend_by_grade, eq_trend = _compute_trend_data(db, period, section)
    db.close()
    return render_template('analysis/trend.html',
                           trend=trend, eq_trend=eq_trend, trend_by_grade=trend_by_grade,
                           period=period, section=section, sections=SECTIONS)


def _compute_trend_data(db, period, section):
    base_sql = "FROM fault_history WHERE 1=1"
    params = []
    if section:
        base_sql += " AND equipment_id IN (SELECT id FROM equipment WHERE section=?)"
        params.append(section)

    label_col = "strftime('%Y-W%W', occurred_at)" if period == 'weekly' else "strftime('%Y-%m', occurred_at)"
    trend = db.execute(f"""
        SELECT {label_col} as label, COUNT(*) as cnt
        {base_sql}
        GROUP BY label ORDER BY label DESC LIMIT 24
    """, params).fetchall()
    trend = list(reversed(trend))

    grade_rows = db.execute(f"""
        SELECT {label_col} as label, COALESCE(grade, 'D') as grade, COUNT(*) as cnt
        {base_sql}
        GROUP BY label, grade
    """, params).fetchall()
    grade_map = {}
    for r in grade_rows:
        grade_map.setdefault(r['label'], {})[r['grade']] = r['cnt']
    trend_by_grade = {
        g: [grade_map.get(t['label'], {}).get(g, 0) for t in trend]
        for g in ('A', 'B', 'C', 'D')
    }

    eq_trend = db.execute(f"""
        SELECT eq_number || ' ' || eq_name as label, COUNT(*) as cnt
        {base_sql}
        GROUP BY equipment_id ORDER BY cnt DESC LIMIT 10
    """, params).fetchall()

    return trend, trend_by_grade, eq_trend


@app.route('/analysis/trend/export')
def analysis_trend_export():
    if not session.get('edit_authorized'):
        flash('엑셀 내보내기는 로그인 후 이용해 주세요.', 'danger')
        return redirect(url_for('login', next=url_for('analysis_trend')))

    period = request.args.get('period', 'monthly')
    section = request.args.get('section', '')
    db = get_db()
    trend, trend_by_grade, eq_trend = _compute_trend_data(db, period, section)
    db.close()

    wb = openpyxl.Workbook()
    period_label = '주별' if period == 'weekly' else '월별'

    ws1 = xlsx_sheet(wb, f'{period_label} 발생추이',
                      ['기간', '건수', 'A 생산정지', 'B 품질영향', 'C 경미고장', 'D 기타'],
                      [14, 10, 12, 12, 12, 10], first=True)
    for i, t in enumerate(trend, 1):
        xlsx_row(ws1, i + 1, [
            t['label'], t['cnt'],
            trend_by_grade['A'][i - 1], trend_by_grade['B'][i - 1],
            trend_by_grade['C'][i - 1], trend_by_grade['D'][i - 1],
        ])

    ws2 = xlsx_sheet(wb, '설비별 고장 TOP10', ['순위', '설비', '건수'], [6, 30, 10])
    for i, row in enumerate(eq_trend, 1):
        xlsx_row(ws2, i + 1, [i, row['label'], row['cnt']])

    fname_prefix = f'추이분석_{period_label}' + (f'_{section}' if section else '')
    log_action('엑셀 내보내기', 'analysis', 0, f'추이 분석 ({period_label}{", " + section if section else ""})')
    return xlsx_response(wb, fname_prefix)


# ──────────────────────────────────────────────
# 엑셀 다운로드
# ──────────────────────────────────────────────
_XLSX_FONT_NAME = '맑은 고딕'
_XLSX_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_XLSX_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name=_XLSX_FONT_NAME)
_XLSX_DATA_FONT = Font(size=10, name=_XLSX_FONT_NAME)
_XLSX_TOTAL_FONT = Font(bold=True, size=10, name=_XLSX_FONT_NAME)
_XLSX_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_XLSX_THIN = Side(style='thin', color='000000')
_XLSX_BORDER = Border(left=_XLSX_THIN, right=_XLSX_THIN, top=_XLSX_THIN, bottom=_XLSX_THIN)
_XLSX_PERCENT_FMT = '0.0"%"'
# 웹 화면과 동일한 브랜드 색상(--mes-*)을 그대로 써서 엑셀 차트와 화면 톤을 맞춘다.
_XLSX_COLOR_DANGER = 'C62828'
_XLSX_COLOR_WARNING = 'E65100'
_XLSX_COLOR_SUCCESS = '2E7D32'
_XLSX_COLOR_NAVY = '0D2137'
_XLSX_COLOR_GRIDLINE = 'D9D9D9'

def xlsx_new_sheet(wb, title, first=False):
    """시트를 만들거나(또는 첫 시트를 재사용) 이름만 붙여서 반환. 내용은 xlsx_title_block/xlsx_table_at으로 채운다."""
    ws = wb.active if first else wb.create_sheet(title)
    if first:
        ws.title = title
    ws.sheet_view.showGridLines = False
    return ws


def xlsx_sheet(wb, title, headers, widths, first=False):
    """기존 방식(표 하나 = 시트 하나)의 단순 리포트용 헬퍼. 헤더=1행 고정."""
    ws = wb.active if first else wb.create_sheet(title)
    if first:
        ws.title = title
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = _XLSX_HEADER_FILL
        cell.font = _XLSX_HEADER_FONT
        cell.alignment = _XLSX_CENTER
        cell.border = _XLSX_BORDER
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    return ws


def xlsx_title_block(ws, num_cols, title, subtitle=None):
    """시트 맨 위에 큰 리포트 제목 + 메타정보(모델/LOT/기간 등) 블록을 쓴다. 다음에 쓸 수 있는 빈 행 번호를 반환."""
    last_col_letter = openpyxl.utils.get_column_letter(max(num_cols, 1))
    ws.merge_cells(f'A1:{last_col_letter}1')
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=16, color='FFFFFF', name=_XLSX_FONT_NAME)
    c.fill = PatternFill("solid", fgColor=_XLSX_COLOR_NAVY)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 32
    row = 2
    if subtitle:
        ws.merge_cells(f'A2:{last_col_letter}2')
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(size=10.5, color='595959', name=_XLSX_FONT_NAME)
        c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c2.fill = PatternFill("solid", fgColor='F2F2F2')
        ws.row_dimensions[2].height = 20
        row = 3
    return row + 1


def _xlsx_estimated_lines(text, width_units):
    """텍스트가 지정한 엑셀 컬럼 폭(width_units) 안에서 줄바꿈되면 몇 줄이 필요한지 추정.
    한글은 라틴 문자의 약 2배 폭을 차지하므로 가중치를 둬서 계산한다."""
    if not text:
        return 1
    text = str(text)
    korean = sum(1 for ch in text if '가' <= ch <= '힣' or 'ㄱ' <= ch <= 'ㆎ')
    effective_len = korean * 2 + (len(text) - korean)
    chars_per_line = max(int(width_units), 4)
    return max(1, -(-effective_len // chars_per_line))  # 올림 나눗셈


def xlsx_kpi_cards(ws, row, ncol, cards):
    """제목 배너 바로 아래에 큰 숫자 KPI 카드를 가로로 나열한다 (전체가동률/총고장건수/총정지시간/MTTR 등 한눈 요약).
    cards: [(label, value_str, color_hex_or_None), ...]. 반환값: 다음에 쓸 수 있는 빈 행 번호."""
    n = len(cards)
    if n == 0:
        return row
    if n > ncol:
        # 카드 수가 시트 폭보다 많으면 span이 0/음수가 되어 merge_cells가 예외를 던진다.
        # 일어나서는 안 되는 상황이지만, 리포트 생성 자체가 죽는 것보단 초과분을 잘라내는 게 안전하다.
        cards = cards[:ncol]
        n = len(cards)

    # 라벨 길이에 비례해서 칸 폭을 배분한다. 예전처럼 균등분할하면 "평균 조치시간(MTTR)"처럼
    # 긴 라벨이 짧은 라벨과 똑같은(혹은 더 좁은) 칸을 받아 글씨가 잘려나가는 문제가 있었다.
    weights = [max(len(c[0]), 4) for c in cards]
    total_w = sum(weights)
    spans = [max(1, round(ncol * w / total_w)) for w in weights]
    diff = ncol - sum(spans)
    if diff != 0:
        idx = spans.index(max(spans))
        spans[idx] = max(1, spans[idx] + diff)

    label_row = row
    value_row = row + 1
    value_row_end = value_row + 1
    card_fill = PatternFill("solid", fgColor='EEF1F5')
    label_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    value_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    max_label_lines = 1
    max_value_lines = 1
    col = 1
    for i, (label, value, color) in enumerate(cards):
        span = spans[i]
        end_col = col + span - 1
        max_label_lines = max(max_label_lines, _xlsx_estimated_lines(label, span * 11))
        # 값이 "95.5%"처럼 짧으면 큰 글씨로, "150분 (12건)"처럼 길어지면 작은 글씨로 —
        # 폰트를 안 줄이면 카드 폭을 넘어가서 옆 카드와 겹쳐 보인다.
        value_len = len(str(value))
        value_font_size = 20 if value_len <= 8 else (16 if value_len <= 13 else 13)
        max_value_lines = max(max_value_lines, _xlsx_estimated_lines(value, span * (11 if value_font_size == 20 else 14)))

        ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=end_col)
        lc = ws.cell(row=label_row, column=col, value=label)
        lc.font = Font(size=9.5, color='6C757D', name=_XLSX_FONT_NAME)
        lc.alignment = label_align
        lc.fill = card_fill

        ws.merge_cells(start_row=value_row, start_column=col, end_row=value_row_end, end_column=end_col)
        vc = ws.cell(row=value_row, column=col, value=value)
        vc.font = Font(size=value_font_size, bold=True, color=(color or _XLSX_COLOR_NAVY), name=_XLSX_FONT_NAME)
        vc.alignment = value_align

        for rr in range(label_row, value_row_end + 1):
            for cc in range(col, end_col + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.border = _XLSX_BORDER
                cell.fill = card_fill
        col = end_col + 1

    ws.row_dimensions[label_row].height = max(16, 12 * max_label_lines + 4)
    value_area_height = max(26, 15 * max_value_lines + 6)
    ws.row_dimensions[value_row].height = value_area_height
    ws.row_dimensions[value_row_end].height = 10
    return value_row_end + 2


def xlsx_block_heading(ws, row, num_cols, text, page_break_before=False):
    """표 블록 사이에 넣는 '▌ 섹션명' 배너. 다음 행(표 헤더가 들어갈 행) 번호를 반환.
    page_break_before=True면 이 배너가 새 인쇄 페이지 맨 위에서 시작하도록 그 앞에 페이지 나누기를 넣는다
    (표가 인쇄 페이지 중간에서 두 동강 나는 걸 막기 위함 — 리포트를 실제로 인쇄/PDF변환할 때 필요)."""
    if page_break_before and row > 1:
        ws.row_breaks.append(Break(id=row - 1))
    last_col_letter = openpyxl.utils.get_column_letter(max(num_cols, 1))
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    c = ws.cell(row=row, column=1, value=f'▌ {text}')
    c.font = Font(bold=True, size=12, color=_XLSX_COLOR_NAVY, name=_XLSX_FONT_NAME)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 22
    return row + 1


def xlsx_subheading(ws, row, num_cols, text):
    """큰 '▌' 블록 안에서 한 단계 더 세분화할 때 쓰는 가벼운 소제목 (예: 섹션별 상세를 섹션 단위로 묶을 때).
    xlsx_block_heading과 달리 페이지를 나누지 않고 색도 옅게 해서, 같은 블록 안의 하위 그룹임을 표시한다."""
    last_col_letter = openpyxl.utils.get_column_letter(max(num_cols, 1))
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    c = ws.cell(row=row, column=1, value=f'· {text}')
    c.font = Font(bold=True, size=10.5, italic=True, color='495057', name=_XLSX_FONT_NAME)
    c.fill = PatternFill("solid", fgColor='F5F6F8')
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = _XLSX_BORDER
    ws.row_dimensions[row].height = 18
    return row + 1


def xlsx_table_at(ws, header_row, headers, widths, rows, percent_cols=None, fill_colors=None, total_cols=None):
    """header_row부터 표 헤더+데이터를 쓴다 (한 시트 안에 여러 표를 세로로 쌓을 때 사용).
    total_cols를 주면(리포트 전체 폭=NCOL), 표의 실제 컬럼 수가 그보다 적을 때 마지막 컬럼을
    total_cols까지 병합해서 늘린다 — 시트를 공유하는 다른 표들과 좌우 폭이 정확히 맞아떨어지게 하기 위함.
    반환값: (header_row, last_data_row, next_free_row)."""
    n = len(headers)
    last_col = max(total_cols or n, n)
    merge_last = last_col > n

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        span_end = last_col if (merge_last and ci == n) else ci
        if span_end > ci:
            ws.merge_cells(start_row=header_row, start_column=ci, end_row=header_row, end_column=span_end)
        for cc in range(ci, span_end + 1):
            cell = ws.cell(row=header_row, column=cc)
            cell.fill = _XLSX_HEADER_FILL
            cell.font = _XLSX_HEADER_FONT
            cell.alignment = _XLSX_CENTER
            cell.border = _XLSX_BORDER
        ws.cell(row=header_row, column=ci, value=h)
        letter = openpyxl.utils.get_column_letter(ci)
        cur = ws.column_dimensions[letter].width
        ws.column_dimensions[letter].width = max(cur or 0, w)
    ws.row_dimensions[header_row].height = 22

    r = header_row + 1
    for i, values in enumerate(rows):
        fill = fill_colors[i] if fill_colors else None
        xlsx_row(ws, r, values, fill_color=fill, percent_cols=percent_cols,
                 merge_last_to=(last_col if merge_last else None), col_widths=widths)
        r += 1
    last_data_row = r - 1
    return header_row, last_data_row, r


def xlsx_row(ws, row_i, values, fill_color=None, percent_cols=None, merge_last_to=None, col_widths=None):
    """한 행을 공통 스타일(중앙정렬+테두리, 선택적 배경색)로 채워 넣는다.
    percent_cols: '95.5' 같은 숫자를 셀에서 '95.5%'로 보이게 할 1-indexed 컬럼들 (값 자체는 그대로 95.5 유지).
    merge_last_to: 주어지면 마지막 값 컬럼을 그 컬럼까지 병합 (표들 간 우측 폭을 맞추기 위함).
    col_widths: 각 컬럼의 엑셀 폭 목록을 주면, 증상/원인처럼 긴 텍스트가 줄바꿈되고도 잘리지 않도록
    실제로 몇 줄이 필요한지 추정해서 행 높이를 자동으로 늘린다 (기존엔 높이가 18로 고정돼 있어
    wrap_text를 켜놔도 셀 안에서 시각적으로 잘려 보였다)."""
    row_fill = PatternFill("solid", fgColor=fill_color) if fill_color else None
    percent_cols = percent_cols or ()
    n = len(values)
    max_lines = 1
    for ci, val in enumerate(values, 1):
        span_end = merge_last_to if (merge_last_to and ci == n and merge_last_to > n) else ci
        if span_end > ci:
            ws.merge_cells(start_row=row_i, start_column=ci, end_row=row_i, end_column=span_end)
        for cc in range(ci, span_end + 1):
            cell = ws.cell(row=row_i, column=cc)
            cell.alignment = _XLSX_CENTER
            cell.border = _XLSX_BORDER
            cell.font = _XLSX_DATA_FONT
            if row_fill:
                cell.fill = row_fill
        cell = ws.cell(row=row_i, column=ci, value=val)
        if ci in percent_cols and isinstance(val, (int, float)):
            cell.number_format = _XLSX_PERCENT_FMT
        if col_widths and ci - 1 < len(col_widths) and isinstance(val, str):
            span_width = sum(col_widths[ci - 1:span_end]) if span_end > ci else col_widths[ci - 1]
            max_lines = max(max_lines, _xlsx_estimated_lines(val, span_width))
    ws.row_dimensions[row_i].height = max(18, 14 * max_lines + 4)

def xlsx_response(wb, filename_prefix):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def xlsx_total_row(ws, first_data_row, last_data_row, label_col, label, value_cols):
    """표 블록 맨 아래에 '합계' 행을 추가하고, 지정한 컬럼들의 합계를 계산해 넣는다.
    first_data_row/last_data_row: 데이터가 있는 첫/마지막 행 번호. value_cols: 합계 낼 1-indexed 컬럼들.
    반환값: (total_row, next_free_row)."""
    total_row = last_data_row + 1
    total_fill = PatternFill("solid", fgColor="E9ECEF")
    cell = ws.cell(row=total_row, column=label_col, value=label)
    cell.font = _XLSX_TOTAL_FONT
    cell.border = _XLSX_BORDER
    cell.fill = total_fill
    cell.alignment = _XLSX_CENTER
    for col in value_cols:
        total = 0
        for r in range(first_data_row, total_row):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                total += v
        tcell = ws.cell(row=total_row, column=col, value=total)
        tcell.font = _XLSX_TOTAL_FONT
        tcell.alignment = _XLSX_CENTER
        tcell.border = _XLSX_BORDER
        tcell.fill = total_fill
    return total_row, total_row + 1


def _xlsx_style_chart_title(chart, size=1300, color=_XLSX_COLOR_NAVY):
    """차트 제목을 진하고 조금 더 크게 (기본값은 작고 가늘어서 보고서에 묻힘)."""
    cp = CharacterProperties(sz=size, b=True, solidFill=color, latin=None)
    try:
        chart.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=cp)
        for r in chart.title.tx.rich.p[0].r:
            r.rPr = cp
    except Exception:
        pass


def xlsx_bar_chart(ws, title, cat_col, data_cols, header_row, last_data_row, start_col, anchor_row=None,
                   y_title='분', x_title='', series_colors=None, point_colors=None, y_percent=False,
                   num_cols=8, min_rows=15):
    """막대그래프를 시트에 삽입. header_row/last_data_row는 '참조할 표 데이터'의 행 범위이고,
    anchor_row는 '차트를 그릴 위치'(생략 시 header_row와 동일) — 같은 표를 참조하는 차트 여러 개를
    세로로 쌓을 때 anchor_row만 바꿔주면 된다.
    표처럼 셀 경계에 정확히 맞춰(두 셀 앵커) 배치하므로 옆 표와 열/행이 가지런히 정렬된다.
    series_colors: data_cols 각각에 대응하는 색상 hex 리스트 (계열이 여럿일 때, 예: 고장=danger/정지=warning).
    point_colors: 단일 계열일 때, 막대 하나하나를 값에 따라 다르게 칠할 hex 리스트 (가동률 임계값 강조용).
    반환값: 차트가 차지하는 마지막 행 번호 (다음 블록/차트를 그 아래로 이어붙일 때 사용)."""
    anchor_row = anchor_row or header_row
    if last_data_row < header_row + 1:
        return anchor_row
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.gapWidth = 60
    chart.overlap = -10
    chart.style = None
    chart.title = title
    _xlsx_style_chart_title(chart)
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=_XLSX_COLOR_GRIDLINE))
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    if y_percent:
        chart.y_axis.numFmt = '0"%"'
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 120  # 막대/라벨이 100%에서 위쪽 테두리·제목과 겹치지 않도록 여유를 더 확보
    # plotArea를 제목 아래로 고정 배치 (막대가 1개뿐인 차트도 제목과 겹치지 않도록 상단 여백을 명시적으로 확보)
    # 주의: openpyxl은 저장 시 chart.plot_area.layout을 chart.layout으로 덮어쓰므로 반드시 chart.layout에 설정해야 한다.
    chart.layout = Layout(
        manualLayout=ManualLayout(x=0.06, y=0.22, h=0.62, w=0.90, xMode="edge", yMode="edge")
    )

    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_data_row)
    for i, dc in enumerate(data_cols):
        data = Reference(ws, min_col=dc, min_row=header_row, max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
        series = chart.series[-1]
        if series_colors and i < len(series_colors):
            series.graphicalProperties.solidFill = series_colors[i]
            series.graphicalProperties.ln = LineProperties(noFill=True)
        if point_colors:
            dpts = []
            for idx, color in enumerate(point_colors):
                dpt = DataPoint(idx=idx)
                dpt.graphicalProperties = GraphicalProperties(solidFill=color)
                dpt.graphicalProperties.ln = LineProperties(noFill=True)
                dpts.append(dpt)
            series.data_points = dpts
    chart.set_categories(cats)

    dl = DataLabelList()
    dl.showVal = True
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showPercent = False
    dl.showBubbleSize = False
    dl.numFmt = '0"%"' if y_percent else '0'
    dl.dLblPos = 'outEnd'
    chart.dataLabels = dl

    # 계열이 하나뿐이면 범례가 제목과 중복돼 자리만 차지하므로 끈다 (여러 계열일 때만 범례 표시)
    if len(data_cols) <= 1:
        chart.legend = None

    table_rows = last_data_row - header_row + 1
    chart_rows = max(table_rows, min_rows)
    end_row = anchor_row + chart_rows - 1
    end_col = start_col + num_cols - 1
    marker_from = AnchorMarker(col=start_col - 1, colOff=0, row=anchor_row - 1, rowOff=0)
    marker_to = AnchorMarker(col=end_col, colOff=0, row=end_row, rowOff=0)
    chart.anchor = TwoCellAnchor(editAs='oneCell', _from=marker_from, to=marker_to)
    ws.add_chart(chart)
    return end_row


_XLSX_PIE_COLORS = ['C62828', 'E65100', '1565C0', '6A1B9A', '00838F', '6C757D', '2E7D32']


def xlsx_hbar_chart(ws, title, cat_col, data_col, header_row, last_data_row, start_col, anchor_row=None,
                    num_cols=8, min_rows=15, point_colors=None, y_percent=True):
    """가로 막대그래프. 항목마다 독립된 한 줄(가로 막대)을 차지하므로, 원형그래프와 달리 항목
    개수나 비중 크기와 무관하게 라벨끼리 절대 겹치지 않는다. 항목별 색상(point_colors) + 막대
    끝의 값 라벨(텍스트) 두 가지로 항목을 동시에 구분한다."""
    if last_data_row < header_row + 1:
        return anchor_row or header_row
    anchor_row = anchor_row or header_row
    chart = BarChart()
    chart.type = "bar"  # 가로 막대: 항목명이 세로로 나란히 나열됨
    chart.grouping = "clustered"
    chart.gapWidth = 50
    chart.style = None
    chart.title = title
    _xlsx_style_chart_title(chart)
    chart.y_axis.title = None
    chart.x_axis.title = None
    # 주의: openpyxl은 막대 방향(bar/col)과 무관하게 x_axis=항목축, y_axis=값축으로 고정한다
    # (barDir="bar"여도 축 객체의 의미는 바뀌지 않고 화면상 방향만 바뀐다).
    chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=_XLSX_COLOR_GRIDLINE))
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    if y_percent:
        chart.y_axis.numFmt = '0"%"'
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
    # 표와 같은 순서(비중이 큰 항목이 위쪽)로 보이도록 항목축 방향을 반전
    chart.x_axis.scaling.orientation = "maxMin"
    # 왼쪽에 긴 한글 항목명이 들어갈 여백을 넉넉히 확보 (제목-플롯 겹침 방지는 y로 확보)
    chart.layout = Layout(
        manualLayout=ManualLayout(x=0.28, y=0.16, h=0.78, w=0.68, xMode="edge", yMode="edge")
    )
    chart.legend = None

    cats = Reference(ws, min_col=cat_col, min_row=header_row + 1, max_row=last_data_row)
    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    series = chart.series[0]
    if point_colors:
        dpts = []
        for idx in range(last_data_row - header_row):
            color = point_colors[idx % len(point_colors)]
            dpt = DataPoint(idx=idx)
            dpt.graphicalProperties = GraphicalProperties(solidFill=color)
            dpt.graphicalProperties.ln = LineProperties(noFill=True)
            dpts.append(dpt)
        series.data_points = dpts

    dl = DataLabelList()
    dl.showVal = True
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showPercent = False
    dl.showBubbleSize = False
    dl.numFmt = '0.0"%"' if y_percent else '0'
    dl.dLblPos = 'outEnd'
    dl.txPr = RichText(p=[Paragraph(
        pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1000, b=True)),
        endParaRPr=CharacterProperties(sz=1000, b=True)
    )])
    chart.dataLabels = dl

    n = last_data_row - header_row
    chart_rows = max(n * 2 + 3, min_rows)  # 항목마다 막대가 뭉개지지 않도록 항목당 최소 2행 확보
    end_row = anchor_row + chart_rows - 1
    end_col = start_col + num_cols - 1
    marker_from = AnchorMarker(col=start_col - 1, colOff=0, row=anchor_row - 1, rowOff=0)
    marker_to = AnchorMarker(col=end_col, colOff=0, row=end_row, rowOff=0)
    chart.anchor = TwoCellAnchor(editAs='oneCell', _from=marker_from, to=marker_to)
    ws.add_chart(chart)
    return end_row


def _xlsx_add_stop_reason_chart(ws, row, ncol, fault_total_min, pause_rows, reason_col, duration_col,
                                page_break_before=False):
    """고장시간 + 정지사유(설비점검/자재대기/교체/기타)별 시간을 집계해서 '정지요인별 분석' 표와
    가로 막대그래프를 그려 넣는다. 설비고장도 결국 정지요인 중 하나이므로 같은 그래프 안에 포함시킨다.
    (이전엔 원형그래프였으나, 정지사유가 6개 이상+소항목 다수일 때 조각 라벨이 서로 겹치는 문제가
    반복적으로 재현되어 가로 막대그래프로 교체 — 항목마다 자기 줄을 가져 구조적으로 겹칠 수 없다.)
    정지시간이 전혀 없으면 아무것도 그리지 않고 그대로 반환."""
    reason_minutes = {}
    for r in pause_rows:
        reason = r[reason_col - 1] or '기타'
        dur = r[duration_col - 1]
        if isinstance(dur, (int, float)):
            reason_minutes[reason] = reason_minutes.get(reason, 0) + dur
    if fault_total_min:
        reason_minutes['설비고장'] = fault_total_min

    total = sum(reason_minutes.values())
    if total <= 0:
        return row

    # 비중이 큰 순서로 정렬해서 어떤 원인이 가장 큰 비중을 차지하는지 표·그래프에서 바로 보이게
    # 한다 (설비 정지원인 분석에서 흔히 쓰는 파레토식 정렬). 막대그래프는 항목이 몇 개든 각자
    # 독립된 줄을 차지하므로, 원형그래프와 달리 항목 수를 줄일 필요가 없다 — 전부 그대로 보여준다.
    ordered_names = sorted(reason_minutes.keys(), key=lambda n: reason_minutes[n], reverse=True)
    table_rows = [[n, reason_minutes[n], round(reason_minutes[n] / total * 100, 1)] for n in ordered_names]

    row = xlsx_block_heading(ws, row, ncol, f'정지요인별 분석 (총 {total}분)', page_break_before=page_break_before)
    hdr, last, row = xlsx_table_at(ws, row, ['정지요인', '시간(분)', '비율(%)'], [16, 12, 12], table_rows,
                                    percent_cols=[3], total_cols=ncol)
    point_colors = [_XLSX_PIE_COLORS[i % len(_XLSX_PIE_COLORS)] for i in range(len(ordered_names))]
    row = xlsx_hbar_chart(ws, '정지요인별 구성비 (비중이 큰 순)', cat_col=1, data_col=3, header_row=hdr, last_data_row=last,
                          start_col=1, anchor_row=row, num_cols=ncol,
                          min_rows=max(15, len(ordered_names) * 2 + 3), point_colors=point_colors) + 1
    return row


def xlsx_finalize_report_sheet(ws, report_title=None, tab_color=None, freeze_row=None, wide=True):
    """여러 표 블록이 세로로 쌓인 '한 시트 종합 리포트' 마무리: 인쇄설정(가로/폭맞춤), 문서 제목·생성일시, 탭 색상.
    freeze_row을 주면 그 행까지(제목 블록) 스크롤해도 항상 보이게 고정."""
    if freeze_row:
        ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    ws.page_setup.orientation = 'landscape' if wide else 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.5, header=0.3, footer=0.3)

    if report_title:
        ws.oddHeader.left.text = report_title
        ws.oddHeader.left.size = 12
        ws.oddHeader.left.font = _XLSX_FONT_NAME + ',Bold'
        ws.oddHeader.right.text = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')} · Sinopex MES"
        ws.oddHeader.right.size = 8
        ws.oddFooter.center.text = "페이지 &P / &N"

    if tab_color:
        ws.sheet_properties.tabColor = tab_color


@app.route('/export/excel')
def export_excel():
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    eq_id     = request.args.get('eq_id', '').strip()
    symptom   = request.args.get('symptom', '').strip()

    sql = "SELECT * FROM fault_history WHERE 1=1"
    params = []
    if date_from:
        sql += " AND date(occurred_at) >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND date(occurred_at) <= ?"
        params.append(date_to)
    if eq_id:
        sql += " AND equipment_id=?"
        params.append(eq_id)
    if symptom:
        sql += " AND symptom=?"
        params.append(symptom)
    sql += " ORDER BY occurred_at DESC"

    db = get_db()
    faults = db.execute(sql, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '고장이력'

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['No', '설비번호', '설비명', '발생일시', '완료일',
               '증상', '원인', '조치내용', '작업자', '고장등급', '사용부품']
    col_widths = [5, 12, 20, 18, 12, 20, 25, 30, 10, 10, 25]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    ws.row_dimensions[1].height = 22

    for ri, fault in enumerate(faults, 2):
        parts_rows = db.execute(
            "SELECT part_name, quantity FROM used_parts WHERE fault_id=?", (fault['id'],)
        ).fetchall()
        parts_str = ', '.join([f"{p['part_name']}x{p['quantity']}" for p in parts_rows])

        row_data = [
            ri - 1,
            fault['eq_number'],
            fault['eq_name'],
            fault['occurred_at'],
            fault['completed_at'] or '',
            fault['symptom'] or '',
            fault['cause'] or '',
            fault['action_detail'] or '',
            fault['worker'] or '',
            fault['grade'] or '',
            parts_str
        ]
        fill_color = 'FFE0E0' if fault['grade'] == 'A' else ('FFFDE0' if fault['grade'] == 'B' else 'E8F5E9')
        row_fill = PatternFill("solid", fgColor=fill_color)

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center
            cell.border = border
            cell.fill = row_fill

    db.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    log_action('엑셀 내보내기', 'fault_history', 0, f'고장이력 ({len(faults)}건)')
    fname = f"고장이력_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ──────────────────────────────────────────────
# 백업
# ──────────────────────────────────────────────
BACKUP_DIR  = os.path.join(BASE_DIR, 'equipment_qr_manager_backup')
BACKUP_LOG  = os.path.join(BASE_DIR, 'backup_log.json')
BACKUP_ITEMS = ['app.py', 'database.py', 'requirements.txt', '실행.bat',
                'equipment_qr.db', 'templates', 'static', 'uploads']
BACKUP_EXCLUDE = {'__pycache__', 'backup_log.json', 'equipment_qr_manager_backup'}

def _ignore_patterns(dir, contents):
    return {c for c in contents if c in BACKUP_EXCLUDE}

def _read_backup_log():
    if os.path.exists(BACKUP_LOG):
        try:
            with open(BACKUP_LOG, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {'last': None}

@app.route('/backup/status')
def backup_status():
    return jsonify(_read_backup_log())

@app.route('/backup', methods=['POST'])
def do_backup():
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        for item in BACKUP_ITEMS:
            src = os.path.join(BASE_DIR, item)
            dst = os.path.join(BACKUP_DIR, item)
            if not os.path.exists(src):
                continue
            if os.path.isdir(src):
                if os.path.exists(dst):
                    shutil.rmtree(dst)
                shutil.copytree(src, dst, ignore=_ignore_patterns)
            else:
                shutil.copy2(src, dst)
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(BACKUP_LOG, 'w', encoding='utf-8') as f:
            json.dump({'last': ts}, f, ensure_ascii=False)
        return jsonify({'ok': True, 'last': ts})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ─── 자동 DB 백업 (매일, 타임스탬프 보관) ───────────────────
AUTO_BACKUP_DIR = os.path.join(BASE_DIR, 'db_backups')
AUTO_BACKUP_RETENTION_DAYS = 14

def _verify_backup_integrity(path):
    """백업 파일이 실제로 열리고 손상되지 않았는지 확인 (PRAGMA integrity_check).
    복사만 하고 한 번도 검증 안 하면, 정작 복원이 필요한 시점에야 손상을 발견하게 된다."""
    if not os.path.exists(path):
        # sqlite3.connect()는 파일이 없으면 자동으로 빈 DB를 새로 만들어버려서,
        # 이 체크 없이는 "백업이 아예 안 됨"이 "정상"으로 오판된다.
        return False
    try:
        conn = sqlite3.connect(path)
        try:
            result = conn.execute('PRAGMA integrity_check').fetchone()
            return bool(result and result[0] == 'ok')
        finally:
            conn.close()
    except Exception:
        return False


def _auto_backup_db():
    os.makedirs(AUTO_BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    src = os.path.join(BASE_DIR, 'equipment_qr.db')
    dst = os.path.join(AUTO_BACKUP_DIR, f'equipment_qr_{ts}.db')
    shutil.copy2(src, dst)

    if _verify_backup_integrity(dst):
        logging.info('자동 DB 백업 완료 및 정합성 확인: %s', dst)
    else:
        logging.error('자동 DB 백업 정합성 검증 실패(손상된 백업): %s', dst)
        if _kakao_connected():
            try:
                _kakao_send_to_all(
                    f'⚠️ 설비QR 자동 백업 정합성 검증 실패\n파일: {os.path.basename(dst)}\n'
                    f'백업 파일이 손상되었을 수 있습니다. 서버 디스크/DB 상태를 확인해주세요.',
                    f'http://{get_local_ip()}:5001/'
                )
            except Exception:
                logging.exception('백업 실패 알림 발송 중 오류')

    cutoff = datetime.now() - timedelta(days=AUTO_BACKUP_RETENTION_DAYS)
    for f in os.listdir(AUTO_BACKUP_DIR):
        p = os.path.join(AUTO_BACKUP_DIR, f)
        if os.path.isfile(p) and datetime.fromtimestamp(os.path.getmtime(p)) < cutoff:
            try:
                os.remove(p)
            except Exception:
                logging.exception('오래된 자동 백업 삭제 실패: %s', p)

def _auto_backup_loop():
    while True:
        try:
            _auto_backup_db()
        except Exception:
            logging.exception('자동 DB 백업 실패')
        time.sleep(24 * 60 * 60)


# ──────────────────────────────────────────────
# 카카오톡 알림 연동
# ──────────────────────────────────────────────
@app.route('/kakao/login')
def kakao_login():
    if not session.get('edit_authorized'):
        flash('카카오 연동은 관리자 인증이 필요합니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(request.referrer or url_for('index'))
    cfg = _kakao_load()
    if not cfg.get('rest_api_key'):
        flash('카카오 REST API 키가 설정되지 않았습니다.', 'danger')
        return redirect(url_for('index'))
    return redirect(_kakao_authorize_url())

@app.route('/kakao/callback')
def kakao_callback():
    error = request.args.get('error')
    code  = request.args.get('code')
    if error or not code:
        flash('카카오 인증 실패: ' + (error or '인증 코드 없음'), 'danger')
        return redirect(url_for('index'))
    result, account = _kakao_exchange_code(code)
    logging.info('kakao_exchange result: %s', result)
    if account:
        flash(f'✅ {account["nickname"]}님, 카카오톡 알림 연동이 완료됐습니다.', 'success')
        try:
            ip = get_local_ip()
            _kakao_send_to_account(account, '✅ 설비 QR 이력관리 시스템 카카오톡 알림이 연결되었습니다.', f'http://{ip}:5001/')
        except Exception:
            logging.exception('kakao welcome message failed')
    else:
        flash('카카오 토큰 발급 실패: ' + str(result.get('error_description', result)), 'danger')
    return redirect(url_for('index'))

@app.route('/kakao/remove/<kakao_id>', methods=['POST'])
def kakao_remove(kakao_id):
    if not session.get('edit_authorized'):
        flash('삭제 권한이 없습니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(request.referrer or url_for('index'))
    _kakao_remove_account(kakao_id)
    flash('카카오톡 알림 대상에서 제거했습니다.', 'success')
    return redirect(url_for('index'))


@app.route('/kakao/notify_settings', methods=['POST'])
def kakao_notify_settings():
    if not session.get('edit_authorized'):
        flash('설정 변경 권한이 없습니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(request.referrer or url_for('index'))
    cfg = {key: (request.form.get(f'notify_{key}') == 'on') for key in NOTIFY_TYPES}
    _save_notify_settings(cfg)
    flash('알림 유형 설정을 저장했습니다.', 'success')
    return redirect(url_for('index'))


# ──────────────────────────────────────────────
# 변경이력 (감사로그)
# ──────────────────────────────────────────────
@app.route('/audit-log')
def audit_log_view():
    if not session.get('edit_authorized'):
        flash('변경이력 조회는 관리자 인증이 필요합니다. 비밀번호를 입력해 주세요.', 'danger')
        return redirect(request.referrer or url_for('index'))
    db = get_db()
    logs = db.execute(
        "SELECT * FROM audit_log ORDER BY id DESC LIMIT 200"
    ).fetchall()
    db.close()
    return render_template('audit_log.html', logs=logs)


# ──────────────────────────────────────────────
# 사진 서빙
# ──────────────────────────────────────────────
@app.route('/uploads/<path:subfolder>/<filename>')
def uploaded_file(subfolder, filename):
    full_path = safe_join(UPLOAD_DIR, subfolder, filename)
    if full_path is None or not os.path.isfile(full_path):
        abort(404)
    return send_file(full_path)


# PWA 서비스워커: 루트 경로(/)에서 서빙해야 하위 경로까지 캐시 제어 가능
@app.route('/sw.js')
def service_worker():
    resp = send_file(os.path.join(BASE_DIR, 'static', 'sw.js'), mimetype='application/javascript')
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


# ──────────────────────────────────────────────
# 에러 페이지
# ──────────────────────────────────────────────
@app.errorhandler(404)
def not_found_error(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    logging.exception('unhandled server error')
    return render_template('errors/500.html'), 500


if __name__ == '__main__':
    init_db()
    threading.Thread(target=_auto_backup_loop, daemon=True).start()
    ip = get_local_ip()
    port = 5001
    print(f"\n{'='*50}")
    print(f"  설비 QR 이력관리 시스템 시작")
    print(f"  로컬 접속 : http://127.0.0.1:{port}")
    print(f"  네트워크  : http://{ip}:{port}")
    print(f"{'='*50}\n")
    try:
        from waitress import serve
        serve(app, host='0.0.0.0', port=port, threads=8, ident=None)
    except ImportError:
        logging.warning('waitress가 설치되어 있지 않아 Flask 개발 서버로 대신 실행합니다. (pip install waitress 권장)')
        app.run(host='0.0.0.0', port=port, debug=False)
